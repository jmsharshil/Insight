import logging
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from core.utils import apply_filters
from django.db.models import Avg, Count, F, Q, ExpressionWrapper, FloatField, Max, Min
from django.db.models.functions import Coalesce
import csv
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from .models import MarkSheet, PublishedResult, RecheckRequest, CheckerQuery
from .serializers import (
    MarkSheetSerializer, PublishedResultSerializer,
    RecheckRequestSerializer, RecheckRequestCreateSerializer,
    RecheckRequestActionSerializer, CheckerQuerySerializer,
    CheckerQueryCreateSerializer,
)
from exams.models import Exam, CheckerToken
from exams.utils import calculate_ranks, generate_checker_token, notify
from exams.emails import send_checker_assignment_email, send_recheck_request_notification

logger = logging.getLogger(__name__)

# ── Role constants ────────────────────────────────────────────────────────────
PAPER_VIEW_ROLES = ['super_admin', 'paper_checker', 'admin_senior_executive']
PAPER_MARK_ROLES = ['super_admin', 'paper_checker']
RECHECK_ROLES = ['super_admin', 'paper_checker', 'admin_senior_executive']
CHECKER_STATUS_ROLES = ['super_admin', 'admin_senior_executive', 'branch_manager']
PUBLISH_ROLES = ['super_admin', 'admin_senior_executive', 'branch_manager']
RECHECK_REQUEST_REVIEW_ROLES = ['super_admin', 'admin_senior_executive', 'branch_manager']
QUERY_ROLES = ['super_admin', 'paper_checker', 'admin_senior_executive']


def _user_role(user):
    return getattr(user, 'role', None)


def build_exam_export_workbook(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Results'
    headers = [
        'Student Name', 'Roll Number', 'Marks Obtained', 'Total Marks',
        'Percentage', 'Rank', 'Is Pass', 'Published At', 'Exam Title'
    ]
    sheet.append(headers)

    red_fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
    green_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')

    for row_index, row in enumerate(rows, start=2):
        sheet.append(row)
        marks_value = row[2]
        try:
            numeric_value = float(marks_value) if marks_value is not None else None
        except (TypeError, ValueError):
            numeric_value = None

        if numeric_value is not None:
            if numeric_value < 30:
                sheet.cell(row=row_index, column=3).fill = red_fill
            elif numeric_value > 60:
                sheet.cell(row=row_index, column=3).fill = green_fill

    for column in sheet.columns:
        non_empty_values = [cell.value for cell in column if cell.value is not None]
        if non_empty_values:
            max_length = max(len(str(value)) for value in non_empty_values)
            sheet.column_dimensions[column[0].column_letter].width = min(max_length + 2, 40)

    return workbook


# ═══════════════════════════════════════════════════════════════════════════════
# 1. GET  /api/v1/exams/{id}/papers/
# ═══════════════════════════════════════════════════════════════════════════════

class PaperView(APIView):
    # permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_submitted', 'is_pass', 'is_rechecked']
    search_fields = ['student__user__name', 'paper_checker__name']
    ordering_fields = '__all__'

    def get(self, request, exam_id):
        role = _user_role(request.user)
        if role not in PAPER_VIEW_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        qs = MarkSheet.objects.filter(exam_id=exam_id).select_related('student__user', 'paper_checker', 'exam__batch', 'exam__subject').prefetch_related('queries')
        if getattr(request.user, 'organization', None):
            qs = qs.filter(exam__branch__organization=request.user.organization)
        if role == 'paper_checker':
            qs = qs.filter(paper_checker=request.user)

        qs = apply_filters(self, request, qs)

        return Response({'success': True, 'count': qs.count(), 'data': MarkSheetSerializer(qs, many=True).data})


# ═══════════════════════════════════════════════════════════════════════════════
# 2. POST  /api/v1/exams/{id}/papers/{marksheet_id}/marks/
# ═══════════════════════════════════════════════════════════════════════════════

class PaperMarksView(APIView):
    # permission_classes = [IsAuthenticated]

    def _get_marksheet(self, request, exam_id, marksheet_id):
        role = _user_role(request.user)
        if role not in PAPER_MARK_ROLES:
            return None, Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            qs = MarkSheet.objects.prefetch_related('queries')
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            ms = qs.get(id=marksheet_id, exam_id=exam_id)
        except MarkSheet.DoesNotExist:
            return None, Response({'success': False, 'message': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if ms.paper_checker != request.user and role != 'super_admin':
            return None, Response({'success': False, 'message': 'Not assigned to you.'}, status=status.HTTP_403_FORBIDDEN)
        # Respect open checker queries for permissions (prevents modification while payroll-excluded)
        if any(q.status == 'open' for q in ms.queries.all()) and role not in ['super_admin', 'admin_senior_executive']:
            return None, Response({
                'success': False,
                'message': 'This marksheet has an open query. Please resolve the query first.',
                'has_open_query': True
            }, status=status.HTTP_403_FORBIDDEN)
        return ms, None

    def post(self, request, exam_id, marksheet_id):
        """POST — submit or update marks on a marksheet. Delegates to PUT for unified logic."""
        return self.put(request, exam_id, marksheet_id)

    def put(self, request, exam_id, marksheet_id):
        """PUT — re-submit / update marks on a marksheet (e.g. after recheck).
        If part of recheck flow, updates RecheckRequest to 'completed', saves checker_notes,
        ensuring the paper is added back to the (new) checker's payroll count.
        """
        ms, err = self._get_marksheet(request, exam_id, marksheet_id)
        if err:
            return err

        marks = request.data.get('marks_obtained')
        if marks is None or float(marks) < 0 or float(marks) > ms.exam.total_marks:
            return Response({'success': False, 'message': 'Invalid marks.'}, status=status.HTTP_400_BAD_REQUEST)

        ms.marks_obtained = marks
        ms.remarks = request.data.get('remarks', '')
        ms.checked_at = timezone.now()
        ms.is_submitted = True
        ms.is_pass = float(marks) >= ms.exam.pass_marks
        ms.save()

        # Handle recheck flow: mark as completed, save notes (fulfills "checker updates ... add notes and resubmits")
        rechecks = RecheckRequest.objects.filter(
            marksheet=ms,
            status__in=['approval_pending', 'approved']
        )
        if rechecks.exists():
            rr = rechecks.latest('created_at')
            rr.status = 'completed'
            rr.checker_notes = request.data.get('notes') or request.data.get('remarks', '')
            rr.save()
            ms.is_rechecked = True
            ms.save(update_fields=['is_rechecked'])

        # Update published result if exists
        try:
            pr = PublishedResult.objects.get(exam_id=exam_id, student=ms.student)
            pr.marks_obtained = marks
            pr.percentage = round((float(marks) / ms.exam.total_marks) * 100, 2) if ms.exam.total_marks else 0
            pr.is_pass = ms.is_pass
            pr.save(update_fields=['marks_obtained', 'percentage', 'is_pass'])
        except PublishedResult.DoesNotExist:
            pass

        return Response({
            'success': True, 
            'message': 'Marks updated. Recheck completed if applicable.',
            'data': MarkSheetSerializer(ms).data
        })

    def delete(self, request, exam_id, marksheet_id):
        """DELETE — remove a marksheet."""
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            qs = MarkSheet.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            ms = qs.get(id=marksheet_id, exam_id=exam_id)
        except MarkSheet.DoesNotExist:
            return Response({'success': False, 'message': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        ms.delete()
        return Response({'success': True, 'message': 'Marksheet deleted.'}, status=status.HTTP_200_OK)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. POST  /api/v1/exams/{id}/papers/{marksheet_id}/recheck/
# ═══════════════════════════════════════════════════════════════════════════════

class PaperRecheckView(APIView):
    # permission_classes = [IsAuthenticated]

    def post(self, request, exam_id, marksheet_id):
        role = _user_role(request.user)
        if role not in RECHECK_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            qs = MarkSheet.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            ms = qs.get(id=marksheet_id, exam_id=exam_id)
        except MarkSheet.DoesNotExist:
            return Response({'success': False, 'message': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        new_cid = request.data.get('new_checker_id')
        if not new_cid:
            return Response({'success': False, 'message': 'new_checker_id required.'}, status=status.HTTP_400_BAD_REQUEST)

        ms.is_rechecked = True
        ms.recheck_request_at = timezone.now()
        ms.paper_checker_id = new_cid
        ms.is_submitted = False
        ms.save()

        generate_checker_token(ms)
        send_checker_assignment_email(ms)

        return Response({'success': True, 'message': 'Recheck requested and reassigned.'})


# ═══════════════════════════════════════════════════════════════════════════════
# 4. GET  /api/v1/exams/{id}/checker-status/
# ═══════════════════════════════════════════════════════════════════════════════

class CheckerStatusView(APIView):
    # permission_classes = [IsAuthenticated]

    def get(self, request, exam_id):
        role = _user_role(request.user)
        if role not in CHECKER_STATUS_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        sheets = MarkSheet.objects.filter(exam_id=exam_id).select_related('paper_checker')
        if getattr(request.user, 'organization', None):
            sheets = sheets.filter(exam__branch__organization=request.user.organization)
        total = sheets.count()
        submitted = sheets.filter(is_submitted=True).count()

        checkers = sheets.values('paper_checker__id', 'paper_checker__name').annotate(
            assigned_count=Count('id'),
            submitted_count=Count('id', filter=Q(is_submitted=True)),
            pending_count=Count('id', filter=Q(is_submitted=False)),
            last_activity=Max('checked_at'),
        )

        res_checkers = [{
            'checker_id': c['paper_checker__id'],
            'checker_name': c['paper_checker__name'],
            'assigned_count': c['assigned_count'],
            'submitted_count': c['submitted_count'],
            'pending_count': c['pending_count'],
            'last_activity': c['last_activity'],
        } for c in checkers if c['paper_checker__id']]

        return Response({
            'success': True,
            'data': {
                'total_papers': total, 'submitted': submitted,
                'approval_pending': total - submitted, 'overdue': 0,
                'checkers': res_checkers,
            },
        })


# ═══════════════════════════════════════════════════════════════════════════════
# 5. POST  /api/v1/checker-portal/submit/  (EXEMPT from auth)
# ═══════════════════════════════════════════════════════════════════════════════

class CheckerPortalSubmitView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        token_str = request.query_params.get('token')
        if not token_str:
            return Response({'success': False, 'message': 'Token required.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            token = CheckerToken.objects.get(token=token_str)
        except CheckerToken.DoesNotExist:
            return Response({'success': False, 'message': 'Invalid token.'}, status=status.HTTP_403_FORBIDDEN)

        if token.is_used or timezone.now() > token.expires_at:
            return Response({'success': False, 'message': 'Token expired or used.'}, status=status.HTTP_403_FORBIDDEN)

        ms = token.marksheet
        marks = request.data.get('marks_obtained')
        if marks is None or float(marks) < 0 or float(marks) > ms.exam.total_marks:
            return Response({'success': False, 'message': 'Invalid marks.'}, status=status.HTTP_400_BAD_REQUEST)

        ms.marks_obtained = marks
        ms.remarks = request.data.get('remarks', '')
        ms.checked_at = timezone.now()
        ms.is_submitted = True
        ms.is_pass = float(marks) >= ms.exam.pass_marks
        ms.save()

        token.is_used = True
        token.save()

        return Response({'success': True, 'message': 'Marks submitted successfully.'})


# ═══════════════════════════════════════════════════════════════════════════════
# 6. POST  /api/v1/exams/{id}/results/publish/
# ═══════════════════════════════════════════════════════════════════════════════

class PublishResultView(APIView):
    # permission_classes = [IsAuthenticated]

    def post(self, request, exam_id):
        role = _user_role(request.user)
        if role not in PUBLISH_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            qs = Exam.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(branch__organization=request.user.organization)
            exam = qs.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        if exam.status == 'results_published':
            return Response({'success': False, 'message': 'Already published.'}, status=status.HTTP_400_BAD_REQUEST)

        if MarkSheet.objects.filter(exam=exam, is_submitted=False).exists():
            return Response({'success': False, 'message': 'Not all marksheets submitted.'}, status=status.HTTP_400_BAD_REQUEST)

        sheets = MarkSheet.objects.filter(exam=exam)
        pubs = []
        for ms in sheets:
            if not PublishedResult.objects.filter(exam=exam, student=ms.student).exists():
                pct = round((float(ms.marks_obtained) / exam.total_marks) * 100, 2) if exam.total_marks else 0
                pubs.append(PublishedResult(
                    exam=exam, student=ms.student, marks_obtained=ms.marks_obtained,
                    total_marks=exam.total_marks, percentage=pct, is_pass=ms.is_pass,
                    published_by=request.user,
                ))
        if pubs:
            PublishedResult.objects.bulk_create(pubs)

        calculate_ranks(exam_id)

        exam.status = 'results_published'
        exam.save()

        top = PublishedResult.objects.filter(exam=exam, rank=1).first()
        top_name = ''
        if top:
            try:
                top_name = top.student.user.name
            except Exception:
                pass

        return Response({
            'success': True, 'message': 'Results published.',
            'data': {'student_count': sheets.count(), 'top_scorer': top_name},
        })


# ═══════════════════════════════════════════════════════════════════════════════
# 7. GET  /api/v1/exams/{id}/results/
# ═══════════════════════════════════════════════════════════════════════════════

class ResultView(APIView):
    # permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_pass']
    search_fields = ['student__user__name']
    ordering_fields = '__all__'

    def get(self, request, exam_id):
        role = _user_role(request.user)
        qs = PublishedResult.objects.filter(exam_id=exam_id).select_related('student__user')
        if getattr(request.user, 'organization', None):
            qs = qs.filter(exam__branch__organization=request.user.organization)

        if role == 'student':
            qs = qs.filter(student__user=request.user)
        elif role == 'parents':
            qs = qs.filter(student__user__linked_parents=request.user)

        qs = apply_filters(self, request, qs)

        return Response({'success': True, 'count': qs.count(), 'data': PublishedResultSerializer(qs, many=True).data})


class ResultDeleteView(APIView):
    """DELETE /api/v1/exams/{exam_id}/results/{result_id}/ — unpublish a result."""
    # permission_classes = [IsAuthenticated]

    def delete(self, request, exam_id, result_id):
        role = _user_role(request.user)
        if role not in PUBLISH_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            qs = PublishedResult.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            pr = qs.get(id=result_id, exam_id=exam_id)
        except PublishedResult.DoesNotExist:
            return Response({'success': False, 'message': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        pr.delete()
        return Response({'success': True, 'message': 'Result deleted.'}, status=status.HTTP_200_OK)


# ═══════════════════════════════════════════════════════════════════════════════
# 8. POST  /api/v1/exams/{id}/results/recheck-request/   (v2 NEW — FRD §4.6.2)
#    GET   /api/v1/exams/{id}/recheck-requests/
#    PATCH /api/v1/exams/{id}/recheck-requests/{request_id}/
# ═══════════════════════════════════════════════════════════════════════════════

class StudentRecheckRequestView(APIView):
    """Student raises a recheck request with reason + optional marksheet upload (per user query).
    Requires Exam.answer_key to be uploaded first. Supports bulk via separate endpoint.
    """
    # permission_classes = [IsAuthenticated]

    def get(self, request, exam_id):
        if _user_role(request.user) != 'student':
            return Response({'success': False, 'message': 'Only students can view their recheck requests.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            from students.models import Student
            student = Student.objects.get(user=request.user)
        except Exception:
            return Response({'success': False, 'message': 'Student profile not found.'}, status=status.HTTP_404_NOT_FOUND)

        rechecks = RecheckRequest.objects.filter(
            marksheet__exam_id=exam_id,
            requested_by=student
        ).select_related('marksheet')
        
        data = RecheckRequestSerializer(rechecks, many=True).data
        # Do not show paper checker info to the student
        for item in data:
            item.pop('new_checker', None)
            item.pop('new_checker_name', None)
            
        return Response({
            'success': True,
            'count': len(data),
            'data': data
        })

    def post(self, request, exam_id):
        if _user_role(request.user) != 'student':
            return Response({'success': False, 'message': 'Only students can request recheck.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            from students.models import Student
            student = Student.objects.get(user=request.user)
        except Exception:
            return Response({'success': False, 'message': 'Student profile not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Must have PublishedResult
        pr_qs = PublishedResult.objects.filter(exam_id=exam_id, student=student)
        if getattr(request.user, 'organization', None):
            pr_qs = pr_qs.filter(exam__branch__organization=request.user.organization)
        if not pr_qs.exists():
            return Response({'success': False, 'message': 'Results not published yet. Cannot request recheck.'}, status=status.HTTP_400_BAD_REQUEST)

        # NEW: Answer key must be uploaded or distributed
        try:
            exam = Exam.objects.get(id=exam_id)
            has_answer_key = bool(exam.answer_key) or exam.answer_key_logs.exists()
            if not has_answer_key:
                return Response({'success': False, 'message': 'Answer key has not been uploaded or distributed yet. Recheck not allowed.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Get marksheet
        try:
            qs = MarkSheet.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            ms = qs.get(exam_id=exam_id, student=student)
        except MarkSheet.DoesNotExist:
            return Response({'success': False, 'message': 'MarkSheet not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Check no pending/approved recheck already exists
        if RecheckRequest.objects.filter(marksheet=ms, status__in=['approval_pending', 'approved']).exists():
            return Response({'success': False, 'message': 'A recheck request is already pending or approved.'}, status=status.HTTP_409_CONFLICT)

        # Support file upload (multipart/form-data)
        serializer_data = request.data.copy()
        uploaded_file = request.FILES.get('uploaded_marksheet')
        if uploaded_file:
            serializer_data['uploaded_marksheet'] = uploaded_file

        ser = RecheckRequestCreateSerializer(data=serializer_data)
        if not ser.is_valid():
            return Response({'success': False, 'errors': ser.errors}, status=status.HTTP_400_BAD_REQUEST)

        rr = RecheckRequest.objects.create(
            marksheet=ms,
            requested_by=student,
            reason=ser.validated_data.get('reason', ''),
            uploaded_marksheet=ser.validated_data.get('uploaded_marksheet'),
            status='approval_pending',
        )

        # Notify ASE
        send_recheck_request_notification(rr)

        return Response({
            'recheck_requested': True, 'status': 'approval_pending',
            'message': 'Your recheck request has been submitted for review.',
            'upload_provided': bool(rr.uploaded_marksheet),
        }, status=status.HTTP_201_CREATED)


        # ser = RecheckRequestActionSerializer(data=request.data)
        # if not ser.is_valid():
        #     return Response({'success': False, 'errors': ser.errors}, status=status.HTTP_400_BAD_REQUEST)

        # action = ser.validated_data['action']
        # ms = rr.marksheet

        # if action == 'approve':
        #     new_checker_id = ser.validated_data['new_checker_id']
        #     rr.status = 'approved'
        #     rr.reviewed_by = request.user
        #     rr.reviewed_at = timezone.now()
        #     rr.new_checker_id = new_checker_id
        #     rr.save()

        #     ms.is_rechecked = True
        #     ms.recheck_request_at = timezone.now()
        #     ms.paper_checker_id = new_checker_id
        #     ms.is_submitted = False
        #     ms.save()

        #     generate_checker_token(ms)
        #     send_checker_assignment_email(ms)

        #     return Response({'success': True, 'message': 'Recheck approved and reassigned to new checker.'})

        # else:  # reject
        #     rr.status = 'rejected'
        #     rr.reviewed_by = request.user
        #     rr.reviewed_at = timezone.now()
        #     rr.save()

        #     # Stub: notify student of rejection
        #     ms = qs.get(id=marksheet_id, exam_id=exam_id)
        # except MarkSheet.DoesNotExist:
        #     return Response({'success': False, 'message': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        # if ms.is_submitted:
        #     return Response({'success': False, 'message': 'Marks already submitted; cannot mark as absent.'}, status=status.HTTP_400_BAD_REQUEST)

        # ms.is_absent = True
        # ms.marks_obtained = 0
        # ms.is_pass = False
        # ms.is_submitted = True
        # ms.remarks = 'Absent'
        # ms.checked_at = timezone.now()
        # ms.save()

        # return Response({'success': True, 'message': 'Student marked as absent.'}, status=status.HTTP_200_OK)


class MarkAllAbsentView(APIView):
    """
    Automatically marks all students who have NO ExamSession record for an exam as absent.
    Should be called after the exam ends.
    """
    # permission_classes = [IsAuthenticated]
    ALLOWED_ROLES = ['super_admin', 'admin_senior_executive', 'branch_manager']

    def post(self, request, exam_id):
        role = _user_role(request.user)
        if role not in self.ALLOWED_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            qs = Exam.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(branch__organization=request.user.organization)
            exam = qs.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        from exams.models import ExamSession

        # Get all students who started an exam session
        attended_student_ids = set(
            ExamSession.objects.filter(exam=exam, is_submitted=True)
            .values_list('student_id', flat=True)
        )

        # Mark absent all marksheets for students without a completed session
        marksheets = MarkSheet.objects.filter(exam=exam, is_submitted=False)
        absent_count = 0
        for ms in marksheets:
            if ms.student_id not in attended_student_ids:
                ms.is_absent = True
                ms.marks_obtained = 0
                ms.is_pass = False
                ms.is_submitted = True
                ms.remarks = 'Absent'
                ms.checked_at = timezone.now()
                ms.save()
                absent_count += 1

        return Response({
            'success': True,
            'message': f'{absent_count} students marked as absent.',
            'absent_count': absent_count,
        }, status=status.HTTP_200_OK)


class BulkRecheckRequestView(APIView):
    """
    NEW: Bulk recheck option for a batch (per user query).
    Creates recheck requests for all students in the exam's batch (if results published).
    Requires answer_key to be uploaded on the Exam first.
    Only ASE/super_admin.
    """
    # permission_classes = [IsAuthenticated]

    def post(self, request, exam_id):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            exam = Exam.objects.get(id=exam_id)
            has_answer_key = bool(getattr(exam, 'answer_key', None)) or exam.answer_key_logs.exists()
            if not has_answer_key:
                return Response({'success': False, 'message': 'Answer key must be uploaded or distributed first for rechecks.'}, status=status.HTTP_400_BAD_REQUEST)
            if not exam.batch:
                return Response({'success': False, 'message': 'Exam has no associated batch for bulk operation.'}, status=status.HTTP_400_BAD_REQUEST)
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        reason = request.data.get('reason', 'Bulk recheck requested for entire batch.')
        from students.models import Student
        # Assume Batch has related students; adjust if Student has batch FK
        students = Student.objects.filter(
            batch=exam.batch,
            user__organization=exam.branch.organization if hasattr(exam.branch, 'organization') else None
        ) if hasattr(Student, 'batch') else []

        created_count = 0
        for student in students:
            try:
                ms = MarkSheet.objects.get(exam=exam, student=student)
                if not RecheckRequest.objects.filter(
                    marksheet=ms, status__in=['approval_pending', 'approved']
                ).exists() and PublishedResult.objects.filter(exam=exam, student=student).exists():
                    RecheckRequest.objects.create(
                        marksheet=ms,
                        requested_by=student,
                        reason=reason,
                        status='approval_pending',
                    )
                    created_count += 1
            except (MarkSheet.DoesNotExist, AttributeError):
                continue

        # Notify ASE about bulk
        logger.info(f'Bulk recheck created {created_count} requests for exam {exam_id}')

        return Response({
            'success': True,
            'message': f'Bulk recheck initiated for batch. {created_count} requests created.',
            'count': created_count,
        }, status=status.HTTP_201_CREATED)


# ═══════════════════════════════════════════════════════════════════════════════
# 11. Paper Checker Query Option (NEW FRD)
# POST  /api/v1/exams/{exam_id}/papers/{marksheet_id}/query/   — raise query (paper_checker)
# PATCH /api/v1/exams/{exam_id}/queries/{query_id}/resolve/     — resolve query (admin/ASE)
# If query raised after recheck (ms.is_rechecked=True), payroll excludes the paper
# from count until query.status == 'resolved'. See CheckerQuery model + updated
# compute_payslip_for_user().
# ═══════════════════════════════════════════════════════════════════════════════

class PaperCheckerQueryView(APIView):
    """Handles raising and resolving paper checker queries.
    Queries prevent payroll payment until completed/resolved.
    """
    # permission_classes = [IsAuthenticated]

    def post(self, request, exam_id, marksheet_id=None):
        """Paper checker raises a query on a marksheet (e.g. answer key missing).
        URL: /exams/{exam_id}/papers/{marksheet_id}/query/
        """
        role = _user_role(request.user)
        if role not in QUERY_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        marksheet_id = marksheet_id or request.data.get('marksheet_id')
        try:
            qs = MarkSheet.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            ms = qs.get(id=marksheet_id, exam_id=exam_id)
        except MarkSheet.DoesNotExist:
            return Response({'success': False, 'message': 'MarkSheet not found.'}, status=status.HTTP_404_NOT_FOUND)

        if ms.paper_checker != request.user and role != 'super_admin':
            return Response({'success': False, 'message': 'Not assigned to you.'}, status=status.HTTP_403_FORBIDDEN)

        # Enforce after recheck start per requirement
        if not getattr(ms, 'is_rechecked', False) and role == 'paper_checker':
            return Response({
                'success': False,
                'message': 'Queries typically raised after recheck starts. Use marks submission instead.'
            }, status=status.HTTP_400_BAD_REQUEST)

        ser = CheckerQueryCreateSerializer(data=request.data)
        if not ser.is_valid():
            return Response({'success': False, 'errors': ser.errors}, status=status.HTTP_400_BAD_REQUEST)

        query = CheckerQuery.objects.create(
            marksheet=ms,
            raised_by=request.user,
            query_type=ser.validated_data['query_type'],
            description=ser.validated_data.get('description', ''),
            status='open',
        )

        # Mark as not submitted until resolved (affects payroll immediately)
        if not ms.is_submitted:
            ms.is_submitted = False
            ms.save(update_fields=['is_submitted'])

        logger.info(f"Checker query raised by {request.user} on marksheet {marksheet_id}: {query.query_type}")

        return Response({
            'success': True,
            'message': 'Query raised. This paper will not count toward payment until resolved.',
            'data': CheckerQuerySerializer(query).data
        }, status=status.HTTP_201_CREATED)

    def patch(self, request, exam_id, query_id=None):
        """Admin/ASE resolves a query. 
        URL: /exams/{exam_id}/queries/{query_id}/resolve/
        Sets status=resolved; payment cut only lifted when completed.
        """
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive']:
            return Response({'success': False, 'message': 'Only admins can resolve queries.'}, status=status.HTTP_403_FORBIDDEN)

        query_id = query_id or request.data.get('query_id')
        try:
            q_qs = CheckerQuery.objects.select_related('marksheet').all()
            if getattr(request.user, 'organization', None):
                q_qs = q_qs.filter(marksheet__exam__branch__organization=request.user.organization)
            query_obj = q_qs.get(id=query_id, marksheet__exam_id=exam_id)
        except (CheckerQuery.DoesNotExist, ValueError):
            return Response({'success': False, 'message': 'Query not found.'}, status=status.HTTP_404_NOT_FOUND)

        if query_obj.status != 'open':
            return Response({'success': False, 'message': 'Query already resolved.'}, status=status.HTTP_400_BAD_REQUEST)

        query_obj.status = 'resolved'
        query_obj.resolved_by = request.user
        query_obj.resolved_at = timezone.now()
        query_obj.save()

        # Optionally allow marks update on resolve (triggers submission for payroll)
        ms = query_obj.marksheet
        if 'marks_obtained' in request.data:
            marks = request.data.get('marks_obtained')
            if marks is not None:
                ms.marks_obtained = marks
                ms.remarks = request.data.get('remarks', ms.remarks or 'Query resolved with marks')
                ms.checked_at = timezone.now()
                ms.is_submitted = True
                ms.is_pass = float(marks) >= (ms.exam.total_marks or 0)
                ms.save()

        logger.info(f"Query {query_obj.id} resolved by {request.user}")

        return Response({
            'success': True,
            'message': 'Query resolved. Paper now eligible for payment on next payroll run (if submitted).',
            'data': CheckerQuerySerializer(query_obj).data
        })


# ═══════════════════════════════════════════════════════════════════════════════
# 9. Mark Absent, Recheck List & Action Views (to satisfy urls.py imports + FRD recheck workflow)
# ═══════════════════════════════════════════════════════════════════════════════

class MarkAbsentView(APIView):
    """POST /api/v1/exams/{exam_id}/papers/{marksheet_id}/mark-absent/ — mark individual student as absent."""

    def post(self, request, exam_id, marksheet_id):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            qs = MarkSheet.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(exam__branch__organization=request.user.organization)
            ms = qs.get(id=marksheet_id, exam_id=exam_id)
        except MarkSheet.DoesNotExist:
            return Response({'success': False, 'message': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if ms.is_submitted:
            return Response({'success': False, 'message': 'Marks already submitted; cannot mark as absent.'}, status=status.HTTP_400_BAD_REQUEST)

        ms.is_absent = True
        ms.marks_obtained = 0
        ms.is_pass = False
        ms.is_submitted = True
        ms.remarks = 'Absent'
        ms.checked_at = timezone.now()
        ms.save()

        return Response({'success': True, 'message': 'Student marked as absent.'}, status=status.HTTP_200_OK)


class RecheckRequestListView(APIView):
    """GET /api/v1/exams/{exam_id}/recheck-requests/ — list all recheck requests for an exam."""

    def get(self, request, exam_id):
        role = _user_role(request.user)
        if role not in RECHECK_REQUEST_REVIEW_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            qs = Exam.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(branch__organization=request.user.organization)
            qs.get(id=exam_id)  # validate exists
        except Exam.DoesNotExist:
            return Response({'success': False, 'message': 'Exam not found.'}, status=status.HTTP_404_NOT_FOUND)

        rechecks = RecheckRequest.objects.filter(
            marksheet__exam_id=exam_id
        ).select_related(
            'marksheet__student__user', 'requested_by', 'reviewed_by', 'new_checker'
        )

        return Response({
            'success': True,
            'count': rechecks.count(),
            'data': RecheckRequestSerializer(rechecks, many=True).data
        })


class RecheckRequestActionView(APIView):
    """PATCH /api/v1/exams/{exam_id}/recheck-requests/{request_id}/ — approve or reject a recheck request by ASE."""

    def patch(self, request, exam_id, request_id):
        role = _user_role(request.user)
        if role not in RECHECK_REQUEST_REVIEW_ROLES:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            qs = RecheckRequest.objects.select_related('marksheet', 'marksheet__exam', 'requested_by')
            if getattr(request.user, 'organization', None):
                qs = qs.filter(marksheet__exam__branch__organization=request.user.organization)
            rr = qs.get(id=request_id, marksheet__exam_id=exam_id)
        except RecheckRequest.DoesNotExist:
            return Response({'success': False, 'message': 'Recheck request not found.'}, status=status.HTTP_404_NOT_FOUND)

        if rr.status != 'approval_pending':
            return Response({'success': False, 'message': f'Can only act on approval_pending requests (current: {rr.status}).'}, status=status.HTTP_400_BAD_REQUEST)

        ser = RecheckRequestActionSerializer(data=request.data)
        if not ser.is_valid():
            return Response({'success': False, 'errors': ser.errors}, status=status.HTTP_400_BAD_REQUEST)

        action = ser.validated_data['action']
        if action == 'approve':
            new_checker_id = ser.validated_data.get('new_checker_id')
            if not new_checker_id:
                return Response({'success': False, 'message': 'new_checker_id is required when approving.'}, status=status.HTTP_400_BAD_REQUEST)

            rr.status = 'approved'
            rr.reviewed_by = request.user
            rr.reviewed_at = timezone.now()
            rr.new_checker_id = new_checker_id
            rr.save()

            ms = rr.marksheet
            ms.is_rechecked = True
            ms.recheck_request_at = timezone.now()
            ms.paper_checker_id = new_checker_id
            ms.is_submitted = False
            ms.save()

            generate_checker_token(ms)
            send_checker_assignment_email(ms)

            # Notify using the helper from exams.utils
            try:
                notify(
                    new_checker_id,
                    title='Paper Recheck Assigned',
                    body=f"You have been assigned a recheck paper for exam: {ms.exam.title}",
                    metadata={"exam_id": str(exam_id), "marksheet_id": str(ms.id), "is_recheck": True},
                )
            except Exception:
                logger.warning('Could not send notification for recheck assignment')

            return Response({'success': True, 'message': 'Recheck approved and reassigned to new checker.'})

        else:  # reject
            rr.status = 'rejected'
            rr.reviewed_by = request.user
            rr.reviewed_at = timezone.now()
            rr.save()

            # Notify student
            try:
                student_id = (getattr(rr.requested_by.user, 'id', None) 
                             if hasattr(rr.requested_by, 'user') and hasattr(rr.requested_by.user, 'id')
                             else getattr(rr.requested_by, 'id', None))
                if student_id:
                    notify(
                        student_id,
                        title='Recheck Request Rejected',
                        body='Your request for recheck has been reviewed and rejected.',
                        metadata={"exam_id": str(exam_id), "recheck_id": str(rr.id)},
                    )
            except Exception:
                logger.warning('Could not send rejection notification')

            return Response({'success': True, 'message': 'Recheck request has been rejected.'})


# ═══════════════════════════════════════════════════════════════════════════════
# 12. Subject-wise / Faculty-wise / Batch / Summary APIs (using PublishedResult + annotations)
# Anchored to Exam.subject, Exam.faculty, Exam.batch relations. No extra models.
# Supports query params: subject_id, faculty_id, batch_id, exam_id
# ═══════════════════════════════════════════════════════════════════════════════

class SubjectWiseResultView(APIView):
    """GET /results/subject-wise/?subject_id=...&batch_id=...&exam_id=...
    Aggregates results by subject (via Exam.subject). Groups ONLY by subject
    to eliminate duplication (overall across exams/batches unless filtered).
    """
    def get(self, request):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive', 'branch_manager']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        qs = PublishedResult.objects.select_related(
            'exam__subject', 'exam__batch', 'exam__branch'
        ).filter(exam__subject__isnull=False)

        if getattr(request.user, 'organization', None):
            qs = qs.filter(exam__branch__organization=request.user.organization)

        # Apply filters (narrows scope before grouping)
        subject_id = request.query_params.get('subject_id')
        batch_id = request.query_params.get('batch_id')
        exam_id = request.query_params.get('exam_id')
        if subject_id:
            qs = qs.filter(exam__subject_id=subject_id)
        if batch_id:
            qs = qs.filter(exam__batch_id=batch_id)
        if exam_id:
            qs = qs.filter(exam_id=exam_id)

        # Pre-annotate for grouping by subject only (no dupes)
        qs = qs.annotate(
            subject_id=F('exam__subject_id'),
            subject_name=F('exam__subject__name'),
        )

        # Aggregate by subject (overall for all exams/subjects in scope)
        annotated = qs.values(
            'subject_id', 'subject_name'
        ).annotate(
            total_students=Count('id'),
            appeared_students=Count('id'),
            passed_students=Count('id', filter=Q(is_pass=True)),
            pass_percentage=ExpressionWrapper(
                F('passed_students') * 100.0 / Coalesce(F('total_students'), 1.0),
                output_field=FloatField()
            ),
            average_marks=Avg('marks_obtained'),
            highest_marks=Max('marks_obtained'),
            lowest_marks=Min('marks_obtained'),
        ).order_by('-pass_percentage', '-average_marks')

        data = list(annotated)
        return Response({
            'success': True,
            'count': len(data),
            'data': data
        })


class FacultyWiseResultView(APIView):
    """GET /results/faculty-wise/?faculty_id=...&subject_id=...&batch_id=...
    Aggregates by faculty ONLY (via Exam.faculty) — one record per faculty
    with overall results across all subjects/batches/exams (unless filtered).
    """
    def get(self, request):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive', 'branch_manager']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        qs = PublishedResult.objects.select_related(
            'exam__faculty__user', 'exam__subject', 'exam__batch', 'exam__branch'
        ).filter(exam__faculty__isnull=False)

        if getattr(request.user, 'organization', None):
            qs = qs.filter(exam__branch__organization=request.user.organization)

        # Apply filters (narrows scope before grouping)
        faculty_id = request.query_params.get('faculty_id')
        subject_id = request.query_params.get('subject_id')
        batch_id = request.query_params.get('batch_id')
        exam_id = request.query_params.get('exam_id')
        if faculty_id:
            qs = qs.filter(exam__faculty_id=faculty_id)
        if subject_id:
            qs = qs.filter(exam__subject_id=subject_id)
        if batch_id:
            qs = qs.filter(exam__batch_id=batch_id)
        if exam_id:
            qs = qs.filter(exam_id=exam_id)

        # Pre-annotate for grouping by faculty only (one record per faculty)
        qs = qs.annotate(
            faculty_id=F('exam__faculty_id'),
            faculty_name=F('exam__faculty__user__name'),
        )

        annotated = qs.values(
            'faculty_id', 'faculty_name'
        ).annotate(
            total_students=Count('id'),
            appeared_students=Count('id'),
            passed_students=Count('id', filter=Q(is_pass=True)),
            pass_percentage=ExpressionWrapper(
                F('passed_students') * 100.0 / Coalesce(F('total_students'), 1.0),
                output_field=FloatField()
            ),
            average_marks=Avg('marks_obtained'),
            highest_marks=Max('marks_obtained'),
            lowest_marks=Min('marks_obtained'),
        ).order_by('-pass_percentage', '-average_marks')

        data = list(annotated)
        return Response({
            'success': True,
            'count': len(data),
            'data': data
        })


class BatchWiseResultView(APIView):
    """GET /results/batch-wise/?batch_id=...&subject_id=...
    Aggregates overall results by batch (via Exam.batch) across ALL exams and subjects.
    """
    def get(self, request):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive', 'branch_manager']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        qs = PublishedResult.objects.select_related(
            'exam__batch', 'exam__subject', 'exam__branch'
        ).filter(exam__batch__isnull=False)

        if getattr(request.user, 'organization', None):
            qs = qs.filter(exam__branch__organization=request.user.organization)

        # Apply filters (narrows scope before grouping)
        batch_id = request.query_params.get('batch_id')
        subject_id = request.query_params.get('subject_id')
        exam_id = request.query_params.get('exam_id')
        if batch_id:
            qs = qs.filter(exam__batch_id=batch_id)
        if subject_id:
            qs = qs.filter(exam__subject_id=subject_id)
        if exam_id:
            qs = qs.filter(exam_id=exam_id)

        # Pre-annotate for grouping by batch only (overall across subjects/exams)
        qs = qs.annotate(
            batch_id=F('exam__batch_id'),
            batch_name=F('exam__batch__name'),
        )

        # Aggregate by batch (overall for all exams and subjects)
        annotated = qs.values(
            'batch_id', 'batch_name'
        ).annotate(
            total_students=Count('id'),
            appeared_students=Count('id'),
            passed_students=Count('id', filter=Q(is_pass=True)),
            pass_percentage=ExpressionWrapper(
                F('passed_students') * 100.0 / Coalesce(F('total_students'), 1.0),
                output_field=FloatField()
            ),
            average_marks=Avg('marks_obtained'),
            highest_marks=Max('marks_obtained'),
            lowest_marks=Min('marks_obtained'),
        ).order_by('-pass_percentage', '-average_marks')

        data = list(annotated)
        return Response({
            'success': True,
            'count': len(data),
            'data': data
        })


class ResultAnalyticsView(APIView):
    """GET /results/analytics/ or /results/summary/ — Overall summary + top performers by category.
    Uses default PublishedResult model with annotations (no extra models).
    """
    def get(self, request):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive', 'branch_manager']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        base_qs = PublishedResult.objects.select_related('exam__subject', 'exam__batch', 'exam__faculty')
        if getattr(request.user, 'organization', None):
            base_qs = base_qs.filter(exam__branch__organization=request.user.organization)

        # Overall summary
        total_results = base_qs.count()
        overall = base_qs.aggregate(
            total_students=Count('id'),
            passed=Count('id', filter=Q(is_pass=True)),
            avg_percentage=Avg('percentage'),
            avg_marks=Avg('marks_obtained'),
        )
        overall_pass_pct = round((overall['passed'] / overall['total_students'] * 100), 2) if overall.get('total_students', 0) > 0 else 0

        # Top 5 subjects by pass %
        subject_summary = base_qs.values(
            'exam__subject__id', 'exam__subject__name'
        ).annotate(
            total=Count('id'),
            passed=Count('id', filter=Q(is_pass=True)),
            pass_pct=ExpressionWrapper(
                F('passed') * 100.0 / Coalesce(F('total'), 1.0), output_field=FloatField()
            ),
            avg_marks=Avg('marks_obtained'),
        ).order_by('-pass_pct')[:5]

        # Top 5 faculty
        faculty_summary = base_qs.values(
            'exam__faculty__id', 'exam__faculty__user__name'
        ).annotate(
            total=Count('id'),
            passed=Count('id', filter=Q(is_pass=True)),
            pass_pct=ExpressionWrapper(
                F('passed') * 100.0 / Coalesce(F('total'), 1.0), output_field=FloatField()
            ),
            avg_marks=Avg('marks_obtained'),
        ).order_by('-pass_pct')[:5]

        # Batch summary
        batch_summary = base_qs.values(
            'exam__batch__id', 'exam__batch__name'
        ).annotate(
            total=Count('id'),
            passed=Count('id', filter=Q(is_pass=True)),
            pass_pct=ExpressionWrapper(
                F('passed') * 100.0 / Coalesce(F('total'), 1.0), output_field=FloatField()
            ),
        ).order_by('-pass_pct')[:5]

        return Response({
            'success': True,
            'data': {
                'overall': {
                    'total_students': overall['total_students'],
                    'passed_students': overall['passed'],
                    'pass_percentage': overall_pass_pct,
                    'average_percentage': round(overall['avg_percentage'] or 0, 2),
                    'average_marks': round(overall['avg_marks'] or 0, 2),
                },
                'top_subjects': list(subject_summary),
                'top_faculty': list(faculty_summary),
                'top_batches': list(batch_summary),
                'total_published_results': total_results,
            }
        })


# ═══════════════════════════════════════════════════════════════════════════════
# 13. Result Export API (CSV download for results or aggregates)
# Supports ?type=exam&exam_id=... or type=subject-wise etc. for analytics integration.
# ═══════════════════════════════════════════════════════════════════════════════

class ResultExportView(APIView):
    """GET /results/export/ — Export results or aggregates as CSV.
    Examples:
      - ?type=exam&exam_id=xxx → per-exam student results CSV
      - ?type=subject-wise&subject_id=... → subject-wise aggregates
      - ?type=analytics → overall summary + top lists (flattened)
    Uses same role checks and scoping as analytics views. No new models.
    """
    def get(self, request):
        role = _user_role(request.user)
        if role not in ['super_admin', 'admin_senior_executive', 'branch_manager']:
            return Response({'success': False, 'message': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)

        export_type = request.query_params.get('type', 'exam')
        exam_id = request.query_params.get('exam_id')
        subject_id = request.query_params.get('subject_id')
        faculty_id = request.query_params.get('faculty_id')
        batch_id = request.query_params.get('batch_id')

        org_filter = getattr(request.user, 'organization', None)

        if export_type == 'exam' and exam_id:
            # Export detailed student results for specific exam (ties to ResultView)
            qs = PublishedResult.objects.filter(exam_id=exam_id).select_related(
                'student__user', 'exam'
            )
            if org_filter:
                qs = qs.filter(exam__branch__organization=org_filter)
            rows = []
            for pr in qs.order_by('rank'):
                student_name = pr.student.user.name if hasattr(pr.student, 'user') and pr.student.user else 'N/A'
                roll_number = getattr(pr.student, 'roll_number', 'N/A') if pr.student else 'N/A'
                marks_value = pr.marks_obtained
                try:
                    marks_value = float(marks_value) if marks_value is not None else None
                except (TypeError, ValueError):
                    marks_value = None
                rows.append([
                    student_name,
                    roll_number,
                    marks_value,
                    pr.total_marks,
                    pr.percentage,
                    getattr(pr, 'rank', 'N/A'),
                    pr.is_pass,
                    pr.published_at,
                    pr.exam.title if pr.exam else ''
                ])

            workbook = build_exam_export_workbook(rows)
            output = BytesIO()
            workbook.save(output)
            response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="results_{export_type}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            return response

        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="results_{export_type}_{timezone.now().strftime("%Y%m%d")}.csv"'
            writer = csv.writer(response)

        if export_type == 'subject-wise':
            # Reuse updated aggregation logic from SubjectWiseResultView (no duplication by subject)
            qs = PublishedResult.objects.select_related(
                'exam__subject', 'exam__batch', 'exam__branch'
            ).filter(exam__subject__isnull=False)
            if org_filter:
                qs = qs.filter(exam__branch__organization=org_filter)
            if subject_id:
                qs = qs.filter(exam__subject_id=subject_id)
            if batch_id:
                qs = qs.filter(exam__batch_id=batch_id)
            if exam_id:
                qs = qs.filter(exam_id=exam_id)

            qs = qs.annotate(
                subject_id=F('exam__subject_id'),
                subject_name=F('exam__subject__name'),
            )
            annotated = qs.values(
                'subject_id', 'subject_name'
            ).annotate(
                total_students=Count('id'),
                passed_students=Count('id', filter=Q(is_pass=True)),
                pass_percentage=ExpressionWrapper(
                    F('passed_students') * 100.0 / Coalesce(F('total_students'), 1.0),
                    output_field=FloatField()
                ),
                average_marks=Avg('marks_obtained'),
                highest_marks=Max('marks_obtained'),
                lowest_marks=Min('marks_obtained'),
            ).order_by('-pass_percentage')

            writer.writerow([
                'Subject ID', 'Subject Name', 'Total Students',
                'Passed Students', 'Pass Percentage', 'Average Marks',
                'Highest Marks', 'Lowest Marks'
            ])
            for row in annotated:
                writer.writerow([
                    row['subject_id'], row['subject_name'],
                    row['total_students'], row['passed_students'],
                    round(row['pass_percentage'], 2), round(row.get('average_marks') or 0, 2),
                    row.get('highest_marks'), row.get('lowest_marks')
                ])
            return response

        elif export_type == 'faculty-wise':
            # Updated to match FacultyWiseResultView: one record per faculty (overall across subjects)
            writer.writerow(['Faculty ID', 'Faculty Name', 'Total Students', 'Passed Students', 'Pass Percentage', 'Average Marks', 'Highest Marks', 'Lowest Marks'])
            qs = PublishedResult.objects.select_related('exam__faculty__user', 'exam__subject', 'exam__batch', 'exam__branch')
            if org_filter:
                qs = qs.filter(exam__branch__organization=org_filter)
            if faculty_id:
                qs = qs.filter(exam__faculty_id=faculty_id)
            if subject_id:
                qs = qs.filter(exam__subject_id=subject_id)
            if batch_id:
                qs = qs.filter(exam__batch_id=batch_id)
            if exam_id:
                qs = qs.filter(exam_id=exam_id)

            qs = qs.annotate(
                faculty_id=F('exam__faculty_id'),
                faculty_name=F('exam__faculty__user__name'),
            )
            annotated = qs.values(
                'faculty_id', 'faculty_name'
            ).annotate(
                total_students=Count('id'),
                passed_students=Count('id', filter=Q(is_pass=True)),
                pass_percentage=ExpressionWrapper(
                    F('passed_students') * 100.0 / Coalesce(F('total_students'), 1.0),
                    output_field=FloatField()
                ),
                average_marks=Avg('marks_obtained'),
                highest_marks=Max('marks_obtained'),
                lowest_marks=Min('marks_obtained'),
            ).order_by('-pass_percentage')

            for row in annotated:
                writer.writerow([
                    row['faculty_id'], row['faculty_name'],
                    row['total_students'], row['passed_students'],
                    round(row['pass_percentage'], 2), round(row.get('average_marks') or 0, 2),
                    row.get('highest_marks'), row.get('lowest_marks')
                ])
            return response

        elif export_type == 'batch-wise':
            writer.writerow(['Batch ID', 'Batch Name', 'Total Students', 'Passed Students', 'Pass Percentage', 'Average Marks', 'Highest Marks', 'Lowest Marks'])
            qs = PublishedResult.objects.select_related('exam__batch', 'exam__branch')
            if org_filter:
                qs = qs.filter(exam__branch__organization=org_filter)
            if batch_id:
                qs = qs.filter(exam__batch_id=batch_id)
            if subject_id:
                qs = qs.filter(exam__subject_id=subject_id)
            if exam_id:
                qs = qs.filter(exam_id=exam_id)

            qs = qs.annotate(
                batch_id=F('exam__batch_id'),
                batch_name=F('exam__batch__name'),
            )
            annotated = qs.values(
                'batch_id', 'batch_name'
            ).annotate(
                total_students=Count('id'),
                passed_students=Count('id', filter=Q(is_pass=True)),
                pass_percentage=ExpressionWrapper(
                    F('passed_students') * 100.0 / Coalesce(F('total_students'), 1.0),
                    output_field=FloatField()
                ),
                average_marks=Avg('marks_obtained'),
                highest_marks=Max('marks_obtained'),
                lowest_marks=Min('marks_obtained'),
            ).order_by('-pass_percentage')

            for row in annotated:
                writer.writerow([
                    row['batch_id'], row['batch_name'],
                    row['total_students'], row['passed_students'],
                    round(row['pass_percentage'], 2), round(row.get('average_marks') or 0, 2),
                    row.get('highest_marks'), row.get('lowest_marks')
                ])
            return response

        elif export_type in ('summary', 'analytics'):
            # Export overall + top lists (flattened with sections)
            base_qs = PublishedResult.objects.select_related('exam__subject', 'exam__batch', 'exam__faculty')
            if org_filter:
                base_qs = base_qs.filter(exam__branch__organization=org_filter)

            overall = base_qs.aggregate(
                total_students=Count('id'),
                passed=Count('id', filter=Q(is_pass=True)),
                avg_percentage=Avg('percentage'),
                avg_marks=Avg('marks_obtained'),
            )
            pass_pct = round((overall['passed'] / overall['total_students'] * 100), 2) if overall.get('total_students', 0) > 0 else 0

            writer.writerow(['Overall Summary'])
            writer.writerow(['Total Students', 'Passed Students', 'Pass Percentage', 'Avg Percentage', 'Avg Marks'])
            writer.writerow([
                overall['total_students'], overall['passed'], pass_pct,
                round(overall['avg_percentage'] or 0, 2), round(overall['avg_marks'] or 0, 2)
            ])
            writer.writerow([])  # separator
            writer.writerow(['Top Subjects'])
            writer.writerow(['Subject Name', 'Total', 'Passed', 'Pass %', 'Avg Marks'])
            subject_summary = base_qs.values('exam__subject__name').annotate(
                total=Count('id'), passed=Count('id', filter=Q(is_pass=True)),
                pass_pct=ExpressionWrapper(F('passed')*100.0/Coalesce(F('total'),1.0), output_field=FloatField()),
                avg_marks=Avg('marks_obtained')
            ).order_by('-pass_pct')[:5]
            for s in subject_summary:
                writer.writerow([s['exam__subject__name'], s['total'], s['passed'], round(s['pass_pct'],2), round(s['avg_marks'] or 0,2)])

            # Add similar rows for top_faculty, top_batches if desired
            writer.writerow([])
            writer.writerow(['Note: Full top-faculty/top-batches included in JSON API. CSV focuses on key metrics.'])
            return response

        return Response({
            'success': False,
            'message': 'Invalid export type. Use: exam, subject-wise, faculty-wise, batch-wise, analytics.'
        }, status=status.HTTP_400_BAD_REQUEST)
