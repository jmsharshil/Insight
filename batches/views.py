import logging
from core.pagination import paginate_queryset

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from core.utils import apply_filters

from django.conf import settings
from django.db import models

from .models import (
    Course, Subject, Batch, BatchStudent, BatchFaculty,
    Classroom, TimetableSlot,
    CourseLevel, Chapter,
)
from .serializers import (
    CourseListSerializer, CourseDetailSerializer, CourseCreateUpdateSerializer,
    SubjectListSerializer, SubjectCreateUpdateSerializer,
    BatchListSerializer, BatchDetailSerializer, BatchCreateUpdateSerializer,
    BatchStudentReadSerializer, AssignStudentsSerializer,
    BatchFacultyReadSerializer, AssignFacultySerializer,
    ClassroomListSerializer, ClassroomCreateUpdateSerializer,
    TimetableSlotListSerializer, TimetableSlotCreateUpdateSerializer,
    FacultyTimetableSerializer, StudentTimetableSerializer,
    CourseLevelSerializer, ChapterSerializer,
)
from .validators import check_faculty_clash, check_classroom_clash, check_batch_clash

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════════════════
#  Course Views
# ═══════════════════════════════════════════════════════════════════════════════

class CourseListView(APIView):
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'code']
    ordering_fields = '__all__'

    def get(self, request):
        # Filter by user's organization
        queryset = Course.objects.prefetch_related('levels__subjects', 'batches').all()
        if getattr(request.user, 'organization', None):
            queryset = queryset.filter(organization=request.user.organization)

        is_active = request.GET.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        queryset = queryset.annotate(
            subject_count=models.Count('levels__subjects', distinct=True)
        )

        queryset = apply_filters(self, request, queryset)

        return paginate_queryset(queryset, request, CourseListSerializer)

    def post(self, request):
        serializer = CourseCreateUpdateSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response(
                {'success': False, 'message': 'Please fix the errors below.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        course = serializer.save()
        return Response(
            {'success': True, 'message': 'Course created successfully.',
             'data': CourseDetailSerializer(course).data},
            status=status.HTTP_201_CREATED,
        )


class CourseDetailView(APIView):

    def _get_course(self, pk):
        try:
            qs = Course.objects.all()
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(organization=self.request.user.organization)
            return qs.get(pk=pk)
        except Course.DoesNotExist:
            return None

    def get(self, request, pk):
        course = self._get_course(pk)
        if course is None:
            return Response({'success': False, 'message': 'Course not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': CourseDetailSerializer(course).data})

    def patch(self, request, pk):
        course = self._get_course(pk)
        if course is None:
            return Response({'success': False, 'message': 'Course not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CourseCreateUpdateSerializer(course, data=request.data, partial=True, context={'request': request})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response({'success': True, 'message': 'Course updated.', 'data': CourseDetailSerializer(course).data})

    def delete(self, request, pk):
        course = self._get_course(pk)
        if course is None:
            return Response({'success': False, 'message': 'Course not found.'}, status=status.HTTP_404_NOT_FOUND)
        course.delete()
        return Response({'success': True, 'message': 'Course deleted.'})


# ═══════════════════════════════════════════════════════════════════════════════
#  CourseLevel Views (E2)
# ═══════════════════════════════════════════════════════════════════════════════

class CourseLevelListView(APIView):
    def get(self, request, course_id):
        levels = CourseLevel.objects.filter(course_id=course_id)
        if getattr(request.user, 'organization', None):
            levels = levels.filter(organization=request.user.organization)
        levels = levels.order_by('order')
        return Response({'success': True, 'data': CourseLevelSerializer(levels, many=True).data})

    def post(self, request, course_id):
        # Ensure course belongs to user's org
        try:
            qs = Course.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(organization=request.user.organization)
            course = qs.get(pk=course_id)
        except Course.DoesNotExist:
            return Response({'success': False, 'message': 'Course not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = CourseLevelSerializer(data=request.data, context={'course': course})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        
        level = serializer.save(course=course, organization=getattr(request.user, 'organization', None))
        return Response({'success': True, 'message': 'Course level created.', 'data': CourseLevelSerializer(level).data}, status=status.HTTP_201_CREATED)


class CourseLevelDetailView(APIView):
    def _get_level(self, course_id, level_id):
        try:
            qs = CourseLevel.objects.filter(course_id=course_id)
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(organization=self.request.user.organization)
            return qs.get(pk=level_id)
        except CourseLevel.DoesNotExist:
            return None

    def get(self, request, course_id, level_id):
        level = self._get_level(course_id, level_id)
        if not level:
            return Response({'success': False, 'message': 'Course level not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': CourseLevelSerializer(level).data})

    def patch(self, request, course_id, level_id):
        level = self._get_level(course_id, level_id)
        if not level:
            return Response({'success': False, 'message': 'Course level not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = CourseLevelSerializer(level, data=request.data, partial=True, context={'course': level.course})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response({'success': True, 'message': 'Course level updated.', 'data': CourseLevelSerializer(level).data})

    def delete(self, request, course_id, level_id):
        level = self._get_level(course_id, level_id)
        if not level:
            return Response({'success': False, 'message': 'Course level not found.'}, status=status.HTTP_404_NOT_FOUND)
        level.delete()
        return Response({'success': True, 'message': 'Course level deleted.'})


# ═══════════════════════════════════════════════════════════════════════════════
#  Subject Views
# ═══════════════════════════════════════════════════════════════════════════════

class SubjectListView(APIView):
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['level', 'level__course', 'is_active']
    search_fields = ['name', 'code']
    ordering_fields = '__all__'

    def get(self, request):
        queryset = Subject.objects.select_related('level', 'level__course').all()
        if getattr(request.user, 'organization', None):
            queryset = queryset.filter(organization=request.user.organization)

        level_id = request.GET.get('level_id')
        course_id = request.GET.get('course_id')
        if level_id:
            queryset = queryset.filter(level_id=level_id)
        if course_id:
            queryset = queryset.filter(level__course_id=course_id)

        is_active = request.GET.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')

        queryset = apply_filters(self, request, queryset)

        return paginate_queryset(queryset, request, SubjectListSerializer)

    def post(self, request):
        serializer = SubjectCreateUpdateSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response(
                {'success': False, 'message': 'Please fix the errors below.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        subject = serializer.save()
        return Response(
            {'success': True, 'message': 'Subject created successfully.',
             'data': SubjectListSerializer(subject).data},
            status=status.HTTP_201_CREATED,
        )


class SubjectDetailView(APIView):

    def _get_subject(self, pk):
        try:
            qs = Subject.objects.all()
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(organization=self.request.user.organization)
            return qs.get(pk=pk)
        except Subject.DoesNotExist:
            return None

    def get(self, request, pk):
        subject = self._get_subject(pk)
        if subject is None:
            return Response({'success': False, 'message': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': SubjectListSerializer(subject).data})

    def patch(self, request, pk):
        subject = self._get_subject(pk)
        if subject is None:
            return Response({'success': False, 'message': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = SubjectCreateUpdateSerializer(subject, data=request.data, partial=True, context={'request': request})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response({'success': True, 'message': 'Subject updated.', 'data': SubjectListSerializer(subject).data})

    def delete(self, request, pk):
        subject = self._get_subject(pk)
        if subject is None:
            return Response({'success': False, 'message': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)
        subject.delete()
        return Response({'success': True, 'message': 'Subject deleted.'})


# ═══════════════════════════════════════════════════════════════════════════════
#  Chapter Views (E2)
# ═══════════════════════════════════════════════════════════════════════════════

class ChapterListView(APIView):
    def get(self, request, subject_id):
        chapters = Chapter.objects.filter(subject_id=subject_id).order_by('order')
        if getattr(request.user, 'organization', None):
            chapters = chapters.filter(subject__organization=request.user.organization)
        return Response({'success': True, 'data': ChapterSerializer(chapters, many=True).data})

    def post(self, request, subject_id):
        try:
            qs = Subject.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(organization=request.user.organization)
            subject = qs.get(pk=subject_id)
        except Subject.DoesNotExist:
            return Response({'success': False, 'message': 'Subject not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = ChapterSerializer(data=request.data, context={'subject': subject})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        
        chapter = serializer.save(subject=subject)
        return Response({'success': True, 'message': 'Chapter created.', 'data': ChapterSerializer(chapter).data}, status=status.HTTP_201_CREATED)


class ChapterDetailView(APIView):
    def _get_chapter(self, subject_id, chapter_id):
        try:
            qs = Chapter.objects.filter(subject_id=subject_id)
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(subject__organization=self.request.user.organization)
            return qs.get(pk=chapter_id)
        except Chapter.DoesNotExist:
            return None

    def get(self, request, subject_id, chapter_id):
        chapter = self._get_chapter(subject_id, chapter_id)
        if not chapter:
            return Response({'success': False, 'message': 'Chapter not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': ChapterSerializer(chapter).data})

    def patch(self, request, subject_id, chapter_id):
        chapter = self._get_chapter(subject_id, chapter_id)
        if not chapter:
            return Response({'success': False, 'message': 'Chapter not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ChapterSerializer(chapter, data=request.data, partial=True, context={'subject': chapter.subject})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response({'success': True, 'message': 'Chapter updated.', 'data': ChapterSerializer(chapter).data})

    def delete(self, request, subject_id, chapter_id):
        chapter = self._get_chapter(subject_id, chapter_id)
        if not chapter:
            return Response({'success': False, 'message': 'Chapter not found.'}, status=status.HTTP_404_NOT_FOUND)
        chapter.delete()
        return Response({'success': True, 'message': 'Chapter deleted.'})

# ═══════════════════════════════════════════════════════════════════════════════
#  Batch Views
# ═══════════════════════════════════════════════════════════════════════════════

class BatchListView(APIView):
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['course', 'is_active', 'branch']
    search_fields = ['name', 'batch_code']
    ordering_fields = '__all__'

    def get(self, request):
        queryset = Batch.objects.select_related('course').prefetch_related('batch_students').all()
        if getattr(request.user, 'organization', None):
            queryset = queryset.filter(organization=request.user.organization)

        course_id = request.GET.get('course_id')
        is_active = request.GET.get('is_active')
        branch_id = request.GET.get('branch_id') or request.GET.get('branch')

        if course_id:
            queryset = queryset.filter(course_id=course_id)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        if branch_id:
            queryset = queryset.filter(branch_id=branch_id)

        # Annotate enrolled student count
        queryset = queryset.annotate(
            enrolled_count=models.Count('batch_students')
        )

        queryset = apply_filters(self, request, queryset)

        return paginate_queryset(queryset, request, BatchListSerializer)

    def post(self, request):
        serializer = BatchCreateUpdateSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response(
                {'success': False, 'message': 'Please fix the errors below.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
        batch = serializer.save()
        batch.refresh_from_db()
        return Response(
            {'success': True, 'message': 'Batch created successfully.',
             'data': BatchDetailSerializer(batch).data},
            status=status.HTTP_201_CREATED,
        )


class BatchDetailView(APIView):

    def _get_batch(self, pk):
        try:
            qs = Batch.objects.select_related('course').all()
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(organization=self.request.user.organization)
            return qs.get(pk=pk)
        except Batch.DoesNotExist:
            return None

    def get(self, request, pk):
        batch = self._get_batch(pk)
        if batch is None:
            return Response({'success': False, 'message': 'Batch not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': BatchDetailSerializer(batch).data})

    def patch(self, request, pk):
        batch = self._get_batch(pk)
        if batch is None:
            return Response({'success': False, 'message': 'Batch not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = BatchCreateUpdateSerializer(batch, data=request.data, partial=True, context={'request': request})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        batch.refresh_from_db()
        return Response({'success': True, 'message': 'Batch updated.', 'data': BatchDetailSerializer(batch).data})

    def delete(self, request, pk):
        batch = self._get_batch(pk)
        if batch is None:
            return Response({'success': False, 'message': 'Batch not found.'}, status=status.HTTP_404_NOT_FOUND)
        batch.delete()
        return Response({'success': True, 'message': 'Batch deleted.'})


# ── Batch Student Assignment ──────────────────────────────────────────────────

from students.models import Student

class BatchAssignStudentsView(APIView):

    def post(self, request, pk):
        try:
            qs = Batch.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(organization=request.user.organization)

            batch = qs.get(pk=pk)

        except Batch.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'message': 'Batch not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = AssignStudentsSerializer(
            data=request.data,
            context={'request': request}
        )

        if not serializer.is_valid():
            return Response(
                {
                    'success': False,
                    'errors': serializer.errors
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        student_ids = serializer.validated_data['student_ids']

        # Fetch Student records
        students = Student.objects.select_related('user').filter(
            id__in=student_ids,
            is_active=True
        )

        found_student_ids = {
            str(student.id)
            for student in students
        }

        requested_student_ids = {
            str(student_id)
            for student_id in student_ids
        }

        invalid_ids = requested_student_ids - found_student_ids

        if invalid_ids:
            return Response(
                {
                    'success': False,
                    'message': f'Invalid student IDs: {list(invalid_ids)}'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        student_ids_list = [
            student.id
            for student in students
        ]

        current_count = BatchStudent.objects.filter(batch=batch).count()
        already_enrolled = set(
            BatchStudent.objects.filter(batch=batch, student_id__in=student_ids_list)
            .values_list('student_id', flat=True)
        )
        to_enroll = [
            student_id
            for student_id in student_ids_list
            if student_id not in already_enrolled
        ]

        if current_count + len(to_enroll) > batch.max_students:
            remaining = batch.max_students - current_count
            return Response(
                {
                    'success': False,
                    'message': f'Batch capacity exceeded. Only {remaining} seats remaining.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        created = []

        for student_id in to_enroll:
            enrollment = BatchStudent.objects.create(
                batch=batch,
                student_id=student_id
            )
            Student.objects.filter(
                id=student_id
            ).update(
                batch=batch,
                current_batch_name=batch.name 
            )
            created.append(enrollment)

        return Response(
            {
                'success': True,
                'message': (
                    f'{len(created)} student(s) enrolled. '
                    f'{len(already_enrolled)} already enrolled.'
                ),
                'data': BatchStudentReadSerializer(
                    created,
                    many=True
                ).data if created else []
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_200_OK
        )

class BatchRemoveStudentView(APIView):

    def post(self, request, pk, student_id):
        try:
            student = Student.objects.select_related('user').get(
                id=student_id
            )
        except Student.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'message': 'Student not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            qs = BatchStudent.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(batch__organization=request.user.organization)

            enrollment = qs.get(batch_id=pk, student_id=student.id)

        except BatchStudent.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'message': 'Student not enrolled in this batch.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        enrollment.delete()
        
         # Clear student's current batch
        student.batch = None
        student.current_batch_name = ''
        student.save(
            update_fields=[
                'batch',
                'current_batch_name',
                'updated_at'
            ]
        )

        return Response(
            {
                'success': True,
                'message': 'Student removed from batch.'
            }
        )

# ── Batch Faculty Assignment ──────────────────────────────────────────────────
from faculty.models import FacultyProfile

class BatchAssignFacultyView(APIView):

    def post(self, request, pk):
        try:
            qs = Batch.objects.all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(organization=request.user.organization)
            batch = qs.get(pk=pk)
        except Batch.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'message': 'Batch not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = AssignFacultySerializer(
            data=request.data,
            context={'request': request}
        )

        if not serializer.is_valid():
            return Response(
                {
                    'success': False,
                    'errors': serializer.errors
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        faculty_id = serializer.validated_data['faculty_id']
        subject_id = serializer.validated_data.get('subject_id')

        try:
            faculty = FacultyProfile.objects.select_related('user').get(
                id=faculty_id,
                is_active=True
            )

        except FacultyProfile.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'message': 'Invalid faculty ID.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        bf, created = BatchFaculty.objects.get_or_create(
            batch=batch,
            faculty_id=faculty.id,
            subject_id=subject_id,
        )

        if not created:
            return Response(
                {
                    'success': False,
                    'message': 'Faculty already assigned to this batch/subject.'
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(
            {
                'success': True,
                'message': 'Faculty assigned successfully.',
                'data': BatchFacultyReadSerializer(bf).data,
            },
            status=status.HTTP_201_CREATED
        )

class BatchRemoveFacultyView(APIView):

    def post(self, request, pk, faculty_id):
        subject_id = request.data.get('subject_id')

        try:
            faculty = FacultyProfile.objects.get(
                id=faculty_id
            )

        except FacultyProfile.DoesNotExist:
            return Response(
                {
                    'success': False,
                    'message': 'Faculty not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        qs = BatchFaculty.objects.filter(batch_id=pk, faculty_id=faculty.id)
        if getattr(request.user, 'organization', None):
            qs = qs.filter(batch__organization=request.user.organization)
        
        if subject_id:
            qs = qs.filter(subject_id=subject_id)

        if not qs.exists():
            return Response(
                {
                    'success': False,
                    'message': 'Faculty assignment not found.'
                },
                status=status.HTTP_404_NOT_FOUND
            )

        qs.delete()

        return Response(
            {
                'success': True,
                'message': 'Faculty removed from batch.'
            }
        )

# ═══════════════════════════════════════════════════════════════════════════════
#  Classroom Views
# ═══════════════════════════════════════════════════════════════════════════════

class ClassroomListView(APIView):
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'building', 'room_number']
    ordering_fields = '__all__'

    def get(self, request):
        queryset = Classroom.objects.all()
        if getattr(request.user, 'organization', None):
            queryset = queryset.filter(organization=request.user.organization)
        is_active = request.GET.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
            
        queryset = apply_filters(self, request, queryset)
        
        serializer = ClassroomListSerializer(queryset, many=True)
        return Response({'success': True, 'count': queryset.count(), 'data': serializer.data})

    def post(self, request):
        serializer = ClassroomCreateUpdateSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        classroom = serializer.save()
        return Response(
            {'success': True, 'message': 'Classroom created.', 'data': ClassroomListSerializer(classroom).data},
            status=status.HTTP_201_CREATED,
        )


class ClassroomDetailView(APIView):

    def _get_classroom(self, pk):
        try:
            qs = Classroom.objects.all()
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(organization=self.request.user.organization)
            return qs.get(pk=pk)
        except Classroom.DoesNotExist:
            return None

    def get(self, request, pk):
        classroom = self._get_classroom(pk)
        if classroom is None:
            return Response({'success': False, 'message': 'Classroom not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': ClassroomListSerializer(classroom).data})

    def patch(self, request, pk):
        classroom = self._get_classroom(pk)
        if classroom is None:
            return Response({'success': False, 'message': 'Classroom not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = ClassroomCreateUpdateSerializer(classroom, data=request.data, partial=True, context={'request': request})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response({'success': True, 'message': 'Classroom updated.', 'data': ClassroomListSerializer(classroom).data})

    def delete(self, request, pk):
        classroom = self._get_classroom(pk)
        if classroom is None:
            return Response({'success': False, 'message': 'Classroom not found.'}, status=status.HTTP_404_NOT_FOUND)
        classroom.delete()
        return Response({'success': True, 'message': 'Classroom deleted.'})



# ═══════════════════════════════════════════════════════════════════════════════
#  Timetable Views
# ═══════════════════════════════════════════════════════════════════════════════

class TimetableListView(APIView):
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    filterset_fields = ['batch', 'day_of_week', 'faculty', 'subject', 'batch__course', 'session_type','batch__branch']
    search_fields = []
    ordering_fields = '__all__'

    def get(self, request):
        queryset = TimetableSlot.objects.select_related(
            'batch', 'batch__course', 'subject', 'faculty', 'classroom', 'batch__branch'
        ).all()
        if getattr(request.user, 'organization', None):
            queryset = queryset.filter(organization=request.user.organization)

        batch_id = request.GET.get('batch_id')
        day_of_week = request.GET.get('day_of_week')
        faculty_id = request.GET.get('faculty_id')
        subject_id = request.GET.get('subject_id')
        course_id = request.GET.get('course_id')
        session_type = request.GET.get('session_type')
        branch_id = request.GET.get('branch_id')

        if batch_id:
            # allow comma-separated batch IDs for E4 filter
            batch_ids = [b.strip() for b in batch_id.split(',')]
            queryset = queryset.filter(batch_id__in=batch_ids)
        if day_of_week is not None:
            queryset = queryset.filter(day_of_week=int(day_of_week))
        if faculty_id:
            queryset = queryset.filter(faculty_id=faculty_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if course_id:
            queryset = queryset.filter(batch__course_id=course_id)
        if session_type:
            queryset = queryset.filter(session_type=session_type)
        if branch_id:
            queryset = queryset.filter(batch__branch_id=branch_id)

        queryset = apply_filters(self, request, queryset)

        serializer = TimetableSlotListSerializer(queryset, many=True)
        return Response({'success': True, 'count': queryset.count(), 'data': serializer.data})

    def post(self, request):
        serializer = TimetableSlotCreateUpdateSerializer(data=request.data, context={'request': request})
        if not serializer.is_valid():
            return Response(
                {'success': False, 'message': 'Please fix the errors below.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data = serializer.validated_data

        day_of_week = data.get('day_of_week')
        session_date = data.get('session_date')
        if day_of_week is None and session_date:
            day_of_week = session_date.weekday()

        start_time = data.get('start_time')
        end_time = data.get('end_time')
        conflicts = []
        all_clash_details = []

        # Clash detection — batch
        if data.get('batch') and start_time and end_time:
            batch_clashes = check_batch_clash(
                batch_id=data['batch'].id,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
            )
            if batch_clashes:
                c = batch_clashes[0]
                conflicts.append(
                    f"Batch '{c['batch_name']}' already has a slot from "
                    f"{c['start_time']}–{c['end_time']} on this day."
                )
                all_clash_details.extend(batch_clashes)

        # Clash detection — faculty
        if data.get('faculty') and start_time and end_time:
            faculty_clashes = check_faculty_clash(
                faculty_id=data['faculty'].id,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
            )
            if faculty_clashes:
                c = faculty_clashes[0]
                conflicts.append(
                    f"Faculty '{c['faculty_name']}' is already scheduled from "
                    f"{c['start_time']}–{c['end_time']} in batch '{c['batch_name']}' on this day."
                )
                all_clash_details.extend(faculty_clashes)

        # Clash detection — classroom
        if data.get('classroom') and start_time and end_time:
            classroom_clashes = check_classroom_clash(
                classroom_id=data['classroom'].id,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
            )
            if classroom_clashes:
                c = classroom_clashes[0]
                conflicts.append(
                    f"Classroom '{c['classroom_name']}' is already booked from "
                    f"{c['start_time']}–{c['end_time']} for batch '{c['batch_name']}' on this day."
                )
                all_clash_details.extend(classroom_clashes)

        if conflicts:
            seen = set()
            unique_clashes = []
            for cl in all_clash_details:
                if cl['id'] not in seen:
                    seen.add(cl['id'])
                    unique_clashes.append(cl)
            return Response(
                {
                    'success': False,
                    'message': ' | '.join(conflicts),
                    'conflicts': conflicts,
                    'clashing_slots': unique_clashes,
                    'can_force': True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        slot = serializer.save(created_by=request.user if request.user.is_authenticated else None)
        return Response(
            {'success': True, 'message': 'Timetable slot created.',
             'data': TimetableSlotListSerializer(slot).data},
            status=status.HTTP_201_CREATED,
        )

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

class TimetableExportExcelView(TimetableListView):
    def get(self, request):
        queryset = TimetableSlot.objects.select_related(
            'batch', 'batch__course', 'subject', 'faculty', 'classroom', 'batch__branch'
        ).all()
        if getattr(request.user, 'organization', None):
            queryset = queryset.filter(organization=request.user.organization)

        # Role-based constraints
        role = getattr(request.user, 'role', None)
        if role == 'student':
            from batches.models import BatchStudent
            try:
                from students.models import Student
                student = Student.objects.get(user=request.user)
                batch_ids = BatchStudent.objects.filter(student=student).values_list('batch_id', flat=True)
                queryset = queryset.filter(batch_id__in=batch_ids)
            except Exception:
                queryset = queryset.none()
        elif role == 'faculty':
            try:
                from faculty.models import FacultyProfile
                faculty = FacultyProfile.objects.get(user=request.user)
                queryset = queryset.filter(faculty_id=faculty.id)
            except Exception:
                queryset = queryset.none()
        elif role in ['parents', 'parent']:
            from batches.models import BatchStudent
            try:
                from students.models import ParentLink
                linked_students = ParentLink.objects.filter(parent=request.user).values_list('student_id', flat=True)
                batch_ids = BatchStudent.objects.filter(student_id__in=linked_students).values_list('batch_id', flat=True)
                queryset = queryset.filter(batch_id__in=batch_ids)
            except Exception:
                queryset = queryset.none()
        elif role not in ['super_admin']:
            # For branch managers and other admins, filter by branch if applicable
            bid = getattr(request.user, 'branch_id', None)
            if not bid and hasattr(request.user, 'profile'):
                bid = getattr(request.user.profile, 'branch_id', None)
            if bid:
                queryset = queryset.filter(batch__branch_id=bid)

        batch_id = request.GET.get('batch_id')
        day_of_week = request.GET.get('day_of_week')
        faculty_id = request.GET.get('faculty_id')
        subject_id = request.GET.get('subject_id')
        course_id = request.GET.get('course_id')
        session_type = request.GET.get('session_type')
        branch_id = request.GET.get('branch_id')

        if batch_id:
            batch_ids = [b.strip() for b in batch_id.split(',')]
            queryset = queryset.filter(batch_id__in=batch_ids)
        if day_of_week is not None:
            queryset = queryset.filter(day_of_week=int(day_of_week))
        if faculty_id:
            queryset = queryset.filter(faculty_id=faculty_id)
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        if course_id:
            queryset = queryset.filter(batch__course_id=course_id)
        if session_type:
            queryset = queryset.filter(session_type=session_type)
        if branch_id:
            queryset = queryset.filter(batch__branch_id=branch_id)

        import datetime
        from django.utils import timezone
        from django.db import models

        today = timezone.now().date()
        date_from_str = request.GET.get('date_from')
        date_to_str = request.GET.get('date_to')

        if not date_from_str and not date_to_str:
            date_from = today - datetime.timedelta(days=today.weekday())
            date_to = date_from + datetime.timedelta(days=6)
        else:
            try:
                date_from = datetime.datetime.strptime(date_from_str, "%Y-%m-%d").date() if date_from_str else None
            except ValueError:
                date_from = None
            try:
                date_to = datetime.datetime.strptime(date_to_str, "%Y-%m-%d").date() if date_to_str else None
            except ValueError:
                date_to = None

        # ── Build Q filter ───────────────────────────────────────────────────
        # For non-recurring (session_date) slots: filter directly on session_date.
        # For recurring slots: use effective_from / effective_to range.
        #   NOTE: slots with effective_from=NULL & effective_to=NULL would match
        #   any date range via the isnull=True OR clauses, so we post-filter
        #   recurring slots below using day_of_week vs. the date window.
        q_filter = models.Q()
        if date_from and date_to:
            q_filter = models.Q(session_date__range=[date_from, date_to]) | (
                models.Q(is_recurring=True) & models.Q(session_date__isnull=True)
            )
        elif date_from:
            q_filter = models.Q(session_date__gte=date_from) | (
                models.Q(is_recurring=True) & models.Q(session_date__isnull=True)
            )
        elif date_to:
            q_filter = models.Q(session_date__lte=date_to) | (
                models.Q(is_recurring=True) & models.Q(session_date__isnull=True)
            )

        if q_filter:
            queryset = queryset.filter(q_filter)

        queryset = apply_filters(self, request, queryset)

        # ── Post-filter: recurring slots whose day_of_week doesn't fall in range ──
        # Recurring slots with effective_from=NULL & effective_to=NULL always pass
        # the DB-level Q filter above, so we must validate day_of_week in Python.
        # Build the set of weekday integers (0=Mon … 6=Sun) present in [date_from, date_to].
        if date_from and date_to:
            days_in_range = set()
            cur = date_from
            while cur <= date_to and len(days_in_range) < 7:
                days_in_range.add(cur.weekday())
                cur += datetime.timedelta(days=1)

            filtered_ids = []
            for slot in queryset:
                if slot.is_recurring and slot.session_date is None:
                    # Only keep if this slot's weekday is actually in the range
                    if slot.day_of_week is not None and slot.day_of_week not in days_in_range:
                        continue
                filtered_ids.append(slot.pk)

            queryset = queryset.filter(pk__in=filtered_ids)
        # ─────────────────────────────────────────────────────────────────────

        # Ensure ordering
        queryset = queryset.order_by('day_of_week', 'start_time')

        # Generate Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Timetable"

        headers = [
            "Batch", "Course", "Branch", "Day of Week", "Date", 
            "Start Time", "End Time", "Subject", "Faculty", 
            "Classroom", "Session Type"
        ]
        
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_num, cell in enumerate(ws[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[cell.column_letter].width = 20
            
        day_map = {
            0: "Monday", 1: "Tuesday", 2: "Wednesday", 
            3: "Thursday", 4: "Friday", 5: "Saturday", 6: "Sunday"
        }
            
        export_rows = []

        # If we have a bounded range, we can expand recurring slots into actual dates.
        # If the range is unbounded, we fall back to 1 row per slot to prevent infinite loops.
        is_bounded = bool(date_from and date_to)

        for slot in queryset:
            batch_name = slot.batch.name if slot.batch else "-"
            course_name = slot.batch.course.name if slot.batch and slot.batch.course else "-"
            branch_name = slot.batch.branch.name if slot.batch and slot.batch.branch else "-"
            day_str = day_map.get(slot.day_of_week, "-") if slot.day_of_week is not None else "-"
            start_str = slot.start_time.strftime("%I:%M %p") if slot.start_time else "-"
            end_str = slot.end_time.strftime("%I:%M %p") if slot.end_time else "-"
            subject_name = slot.subject.name if slot.subject else "-"
            faculty_name = slot.faculty.user.name if slot.faculty and hasattr(slot.faculty, 'user') else "-"
            classroom_name = slot.classroom.name if slot.classroom else "-"
            session_type = slot.get_session_type_display() if slot.session_type else "-"

            base_row = [
                batch_name, course_name, branch_name, day_str,
                "-", # placeholder for date
                start_str, end_str, subject_name, faculty_name,
                classroom_name, session_type
            ]
            
            # Use datetime.time.min for sorting if start_time is None
            sort_time = slot.start_time if slot.start_time else datetime.time.min

            if slot.session_date:
                # Non-recurring: one exact date
                row = list(base_row)
                row[4] = slot.session_date.strftime("%Y-%m-%d")
                export_rows.append((slot.session_date, sort_time, row))
            elif slot.is_recurring and slot.day_of_week is not None and is_bounded:
                # Recurring bounded: generate a row for EVERY matching weekday in the range
                curr = date_from
                while curr <= date_to:
                    if curr.weekday() == slot.day_of_week:
                        row = list(base_row)
                        row[4] = curr.strftime("%Y-%m-%d")
                        export_rows.append((curr, sort_time, row))
                    curr += datetime.timedelta(days=1)
            else:
                # Fallback for unbounded recurring slots (just output one generic row)
                row = list(base_row)
                row[4] = "Recurring (Every " + day_str + ")"
                export_rows.append((datetime.date.min, sort_time, row))

        # Sort all rows chronologically by Date (x[0]), then Time (x[1])
        export_rows.sort(key=lambda x: (x[0], x[1]))

        for _, _, row in export_rows:
            ws.append(row)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Timetable_Export.xlsx"'
        wb.save(response)
        return response

class TimetableDetailView(APIView):

    def _get_slot(self, pk):
        try:
            qs = TimetableSlot.objects.select_related(
                'batch', 'subject', 'faculty', 'classroom'
            ).all()
            if getattr(self.request.user, 'organization', None):
                qs = qs.filter(organization=self.request.user.organization)
            return qs.get(pk=pk)
        except TimetableSlot.DoesNotExist:
            return None

    def get(self, request, pk):
        slot = self._get_slot(pk)
        if slot is None:
            return Response({'success': False, 'message': 'Timetable slot not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'success': True, 'data': TimetableSlotListSerializer(slot).data})

    def patch(self, request, pk):
        slot = self._get_slot(pk)
        if slot is None:
            return Response({'success': False, 'message': 'Timetable slot not found.'}, status=status.HTTP_404_NOT_FOUND)

        serializer = TimetableSlotCreateUpdateSerializer(slot, data=request.data, partial=True, context={'request': request})
        if not serializer.is_valid():
            return Response({'success': False, 'errors': serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        batch = data.get('batch', slot.batch)
        faculty = data.get('faculty', slot.faculty)
        classroom = data.get('classroom', slot.classroom)
        day = data.get('day_of_week', slot.day_of_week)
        session_date = data.get('session_date', slot.session_date)
        if day is None and session_date:
            day = session_date.weekday()
        start = data.get('start_time', slot.start_time)
        end = data.get('end_time', slot.end_time)

        conflicts = []
        all_clash_details = []

        # Re-run clash detection — batch
        if batch and start and end:
            batch_clashes = check_batch_clash(
                batch_id=batch.id, day_of_week=day,
                start_time=start, end_time=end, exclude_id=slot.id,
            )
            if batch_clashes:
                c = batch_clashes[0]
                conflicts.append(
                    f"Batch '{c['batch_name']}' already has a slot from "
                    f"{c['start_time']}–{c['end_time']} on this day."
                )
                all_clash_details.extend(batch_clashes)

        # Re-run clash detection — faculty
        if faculty and start and end:
            faculty_clashes = check_faculty_clash(
                faculty_id=faculty.id, day_of_week=day,
                start_time=start, end_time=end, exclude_id=slot.id,
            )
            if faculty_clashes:
                c = faculty_clashes[0]
                conflicts.append(
                    f"Faculty '{c['faculty_name']}' is already scheduled from "
                    f"{c['start_time']}–{c['end_time']} in batch '{c['batch_name']}' on this day."
                )
                all_clash_details.extend(faculty_clashes)

        # Re-run clash detection — classroom
        if classroom and start and end:
            classroom_clashes = check_classroom_clash(
                classroom_id=classroom.id, day_of_week=day,
                start_time=start, end_time=end, exclude_id=slot.id,
            )
            if classroom_clashes:
                c = classroom_clashes[0]
                conflicts.append(
                    f"Classroom '{c['classroom_name']}' is already booked from "
                    f"{c['start_time']}–{c['end_time']} for batch '{c['batch_name']}' on this day."
                )
                all_clash_details.extend(classroom_clashes)

        if conflicts:
            seen = set()
            unique_clashes = []
            for cl in all_clash_details:
                if cl['id'] not in seen:
                    seen.add(cl['id'])
                    unique_clashes.append(cl)
            return Response(
                {
                    'success': False,
                    'message': ' | '.join(conflicts),
                    'conflicts': conflicts,
                    'clashing_slots': unique_clashes,
                    'can_force': True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        serializer.save()
        return Response({'success': True, 'message': 'Timetable slot updated.',
                         'data': TimetableSlotListSerializer(slot).data})

    def delete(self, request, pk):
        slot = self._get_slot(pk)
        if slot is None:
            return Response({'success': False, 'message': 'Timetable slot not found.'}, status=status.HTTP_404_NOT_FOUND)
        slot.delete()
        return Response({'success': True, 'message': 'Timetable slot deleted.'})


class TimetableDuplicateSlotView(APIView):
    """
    POST /api/v1/timetable/<uuid:pk>/duplicate/
    Duplicates a regular-session timetable slot to a new day + slot code.

    Body: { "slot_code": "P2", "day_of_week": 3 }
    """

    def post(self, request, pk):
        # 1. Fetch the source slot
        try:
            qs = TimetableSlot.objects.select_related(
                'batch', 'subject', 'faculty', 'classroom'
            ).all()
            if getattr(request.user, 'organization', None):
                qs = qs.filter(organization=request.user.organization)
            source = qs.get(pk=pk)
        except TimetableSlot.DoesNotExist:
            return Response(
                {'success': False, 'message': 'Source timetable slot not found.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # 2. Only regular sessions may be duplicated
        if source.session_type != 'regular':
            return Response(
                {'success': False, 'message': 'Only regular sessions can be duplicated.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 3. Validate inputs
        slot_code = request.data.get('slot_code')
        day_of_week = request.data.get('day_of_week')
        session_date = request.data.get('session_date')

        from batches.constants import FIXED_SLOTS
        from batches.models import SLOT_CODE_CHOICES, DAY_CHOICES

        valid_codes = [c[0] for c in SLOT_CODE_CHOICES]
        valid_days  = [d[0] for d in DAY_CHOICES]

        if not slot_code or slot_code not in valid_codes:
            return Response(
                {'success': False, 'message': f"Invalid slot_code. Choose from {valid_codes}."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if day_of_week is None and session_date is None:
            return Response(
                {'success': False, 'message': 'day_of_week or session_date is required.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if day_of_week is not None:
            try:
                day_of_week = int(day_of_week)
            except (ValueError, TypeError):
                return Response(
                    {'success': False, 'message': 'day_of_week must be an integer.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            if day_of_week not in valid_days:
                return Response(
                    {'success': False, 'message': f"Invalid day_of_week. Choose from {valid_days}."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        elif session_date:
            from datetime import date
            if isinstance(session_date, str):
                try:
                    parsed_date = date.fromisoformat(session_date)
                    day_of_week = parsed_date.weekday()
                    session_date = parsed_date
                except ValueError:
                    return Response({'success': False, 'message': 'Invalid session_date format.'}, status=400)
            else:
                day_of_week = session_date.weekday()

        # 4. Resolve start/end times from slot code
        if slot_code in ['P5', 'P6']:
            from datetime import time
            start_time_str = request.data.get('start_time')
            end_time_str = request.data.get('end_time')
            if not start_time_str or not end_time_str:
                return Response(
                    {'success': False, 'message': 'start_time and end_time required for P5/P6.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            try:
                start_time = time.fromisoformat(start_time_str)
                end_time = time.fromisoformat(end_time_str)
            except ValueError:
                return Response({'success': False, 'message': 'Invalid time format.'}, status=400)
                
            if start_time >= end_time:
                return Response(
                    {'success': False, 'message': 'end_time must be after start_time.'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
        else:
            start_time, end_time = FIXED_SLOTS[slot_code]

        # 5. Clash detection — batch, faculty, classroom
        conflicts = []
        all_clash_details = []

        if source.batch_id:
            batch_clashes = check_batch_clash(
                batch_id=source.batch_id,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
            )
            if batch_clashes:
                c = batch_clashes[0]
                conflicts.append(
                    f"Batch '{c['batch_name']}' already has a slot from "
                    f"{c['start_time']}–{c['end_time']} on this day."
                )
                all_clash_details.extend(batch_clashes)

        if source.faculty_id:
            faculty_clashes = check_faculty_clash(
                faculty_id=source.faculty_id,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
            )
            if faculty_clashes:
                c = faculty_clashes[0]
                conflicts.append(
                    f"Faculty '{c['faculty_name']}' is already scheduled from "
                    f"{c['start_time']}–{c['end_time']} in batch '{c['batch_name']}' on this day."
                )
                all_clash_details.extend(faculty_clashes)

        # 6. Clash detection — classroom
        if source.classroom_id:
            classroom_clashes = check_classroom_clash(
                classroom_id=source.classroom_id,
                day_of_week=day_of_week,
                start_time=start_time,
                end_time=end_time,
            )
            if classroom_clashes:
                c = classroom_clashes[0]
                conflicts.append(
                    f"Classroom '{c['classroom_name']}' is already booked from "
                    f"{c['start_time']}–{c['end_time']} for batch '{c['batch_name']}' on this day."
                )
                all_clash_details.extend(classroom_clashes)

        if conflicts:
            seen = set()
            unique_clashes = []
            for cl in all_clash_details:
                if cl['id'] not in seen:
                    seen.add(cl['id'])
                    unique_clashes.append(cl)
            return Response(
                {
                    'success': False,
                    'message': ' | '.join(conflicts),
                    'conflicts': conflicts,
                    'clashing_slots': unique_clashes,
                    'can_force': True,
                },
                status=status.HTTP_409_CONFLICT,
            )

        # 7. Create the duplicate slot
        new_slot = TimetableSlot.objects.create(
            organization=source.organization,
            batch=source.batch,
            subject=source.subject,
            faculty=source.faculty,
            classroom=source.classroom,
            day_of_week=day_of_week,
            start_time=start_time,
            end_time=end_time,
            is_recurring=source.is_recurring,
            effective_from=source.effective_from,
            effective_to=source.effective_to,
            session_type='regular',
            session_name=source.session_name,
            slot_code=slot_code,
            created_by=request.user if request.user.is_authenticated else None,
        )

        return Response(
            {'success': True, 'message': 'Timetable slot duplicated.',
             'data': TimetableSlotListSerializer(new_slot).data},
            status=status.HTTP_201_CREATED,
        )


# ── Personal Timetable Views ─────────────────────────────────────────────────

# ── Force-create (confirm) View ─────────────────────────────────────────────

class TimetableConfirmView(APIView):
    """
    POST /api/v1/timetable/confirm/
    Force-creates a timetable slot, bypassing all faculty / classroom / batch
    clash checks.  Accepts the same payload as TimetableListView.post.
    """

    def post(self, request):
        serializer = TimetableSlotCreateUpdateSerializer(
            data=request.data, context={'request': request}
        )
        if not serializer.is_valid():
            return Response(
                {'success': False, 'message': 'Please fix the errors below.', 'errors': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )

        slot = serializer.save(
            created_by=request.user if request.user.is_authenticated else None
        )
        return Response(
            {
                'success': True,
                'message': 'Timetable slot force-created (conflicts ignored).',
                'data': TimetableSlotListSerializer(slot).data,
            },
            status=status.HTTP_201_CREATED,
        )


class FacultyTimetableView(APIView):
    """GET /api/v1/timetable/faculty/<faculty_id>/ — weekly schedule for a faculty member."""

    def get(self, request, faculty_id):
        slots = TimetableSlot.objects.select_related(
            'batch', 'subject', 'classroom'
        ).filter(faculty_id=faculty_id)
        if getattr(request.user, 'organization', None):
            slots = slots.filter(organization=request.user.organization)
        slots = slots.order_by('day_of_week', 'start_time')

        serializer = FacultyTimetableSerializer(slots, many=True)

        # Group by day
        grouped = {}
        for s in serializer.data:
            day = s['day_label']
            grouped.setdefault(day, []).append(s)

        return Response({'success': True, 'data': grouped})


class StudentTimetableView(APIView):
    """GET /api/v1/timetable/student/<student_id>/ — weekly schedule for a student."""

    def get(self, request, student_id):
        # Find all batches the student is enrolled in
        batch_ids = BatchStudent.objects.filter(
            student_id=student_id
        ).values_list('batch_id', flat=True)

        slots = TimetableSlot.objects.select_related(
            'subject', 'faculty', 'classroom'
        ).filter(batch_id__in=batch_ids)
        if getattr(request.user, 'organization', None):
            slots = slots.filter(organization=request.user.organization)
        slots = slots.order_by('day_of_week', 'start_time')

        serializer = StudentTimetableSerializer(slots, many=True)

        grouped = {}
        for s in serializer.data:
            day = s['day_label']
            grouped.setdefault(day, []).append(s)

        return Response({'success': True, 'data': grouped})


from branch.models import Branch

class AcademicDropdownsView(APIView):
    """
    GET /api/v1/batches/dropdowns/
    Returns minimal dropdown data for course, level, batch, subject, branch, and classroom.
    """
    def get(self, request):
        user = request.user
        org_id = user.organization_id if hasattr(user, 'organization_id') and user.organization_id else None

        courses_qs = Course.objects.all()
        levels_qs = CourseLevel.objects.all()
        batches_qs = Batch.objects.all()
        subjects_qs = Subject.objects.all()
        branches_qs = Branch.objects.all()
        classrooms_qs = Classroom.objects.all()
        chapters_qs = Chapter.objects.all()

        from exams.models import SubjectPaper
        papers_qs = SubjectPaper.objects.all()

        if org_id:
            courses_qs = courses_qs.filter(organization_id=org_id)
            levels_qs = levels_qs.filter(organization_id=org_id)
            batches_qs = batches_qs.filter(organization_id=org_id)
            subjects_qs = subjects_qs.filter(organization_id=org_id)
            branches_qs = branches_qs.filter(organization_id=org_id)
            classrooms_qs = classrooms_qs.filter(organization_id=org_id)
            chapters_qs = chapters_qs.filter(subject__organization_id=org_id)
            papers_qs = papers_qs.filter(subject__organization_id=org_id)
        
        branch_id = request.GET.get('branch_id')
        if branch_id:
            batches_qs = batches_qs.filter(branch_id=branch_id)
            branches_qs = branches_qs.filter(id=branch_id)
 
        subjects = list(subjects_qs.values('id', 'name', 'level_id'))
        chapters = list(chapters_qs.values('id', 'name', 'subject_id', 'order'))
        papers = list(papers_qs.values('id', 'set_name', 'subject_id', 'file', 'answer_key'))

        chapters_by_subject = {}
        for chapter in chapters:
            subj_id = chapter['subject_id']
            if subj_id not in chapters_by_subject:
                chapters_by_subject[subj_id] = []
            chapters_by_subject[subj_id].append({
                'id': chapter['id'],
                'name': chapter['name'],
                'order': chapter['order']
            })
            
        papers_by_subject = {}
        for paper in papers:
            subj_id = paper['subject_id']
            if subj_id not in papers_by_subject:
                papers_by_subject[subj_id] = []
            papers_by_subject[subj_id].append({
                'id': paper['id'],
                'set_name': paper['set_name'],
                'file': paper['file'],
                'answer_key': paper['answer_key']
            })
        
        for subject in subjects:
            subject['chapters'] = chapters_by_subject.get(subject['id'], [])
            subject['papers'] = papers_by_subject.get(subject['id'], [])

        return Response({
            "success": True,
            "data": {
                "courses": list(courses_qs.values('id', 'name')),
                "levels": list(levels_qs.values('id', 'name', 'course_id')),
                "batches": list(batches_qs.values('id', 'name', 'course_id')),
                "subjects": subjects,
                "branches": list(branches_qs.values('id', 'name', 'city')),
                "classrooms": list(classrooms_qs.values('id', 'name', 'capacity')),
            }
        })


class TimetablePublishView(APIView):
    """
    POST /api/v1/batches/timetable/publish/
    Notifies faculty, students, and their parents that the timetable has been updated.
    Payload: {"batch_id": "<uuid>"} or {"branch_id": "<uuid>"}
    """
    def post(self, request):
        from .models import TimetableSlot, BatchStudent
        from students.models import ParentLink
        from chat.notifications import send_system_notification
        
        batch_id = request.data.get('batch_id')
        branch_id = request.data.get('branch_id')
        
        if not batch_id and not branch_id:
            return Response({'success': False, 'message': 'Please provide batch_id or branch_id.'}, status=status.HTTP_400_BAD_REQUEST)
            
        slots = TimetableSlot.objects.all()
        if batch_id:
            slots = slots.filter(batch_id=batch_id)
        if branch_id:
            slots = slots.filter(batch__branch_id=branch_id)
            
        if getattr(request.user, 'organization', None):
            slots = slots.filter(organization=request.user.organization)
            
        if not slots.exists():
            return Response({'success': False, 'message': 'No timetable slots found for the given criteria.'}, status=status.HTTP_404_NOT_FOUND)
            
        # Get unique faculty members
        faculty_ids = set(slots.values_list('faculty__user_id', flat=True).exclude(faculty__user_id__isnull=True))
        
        # Get unique batches
        batch_ids = set(slots.values_list('batch_id', flat=True))
        
        # Get students in those batches
        batch_students = BatchStudent.objects.filter(batch_id__in=batch_ids).select_related('student__user')
        student_user_ids = set()
        student_profile_ids = set()
        
        for bs in batch_students:
            if getattr(bs.student, 'user', None):
                student_user_ids.add(bs.student.user.id)
            student_profile_ids.add(bs.student.id)
            
        # Get parents for those students
        parents = ParentLink.objects.filter(student_id__in=student_profile_ids).select_related('parent')
        parent_user_ids = set(p.parent.id for p in parents if p.parent)
        
        # Send notifications
        title = "Timetable Updated"
        body_faculty = "Your timetable has been updated. Please check your schedule."
        body_student = "The timetable for your batch has been updated. Please check."
        body_parent = "The timetable for your child has been updated."
        
        notified_count = 0
        for f_id in faculty_ids:
            send_system_notification(str(f_id), title, body_faculty, {"type": "timetable_publish"})
            notified_count += 1
            
        for s_id in student_user_ids:
            send_system_notification(str(s_id), title, body_student, {"type": "timetable_publish"})
            notified_count += 1
            
        for p_id in parent_user_ids:
            send_system_notification(str(p_id), title, body_parent, {"type": "timetable_publish"})
            notified_count += 1
            
        return Response({
            'success': True,
            'message': f'Timetable published successfully. Notified {notified_count} users.',
        })
