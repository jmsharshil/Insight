"""
Management command: import_chapters
────────────────────────────────────
Reads the "CHAPTER, CLASS TEST, PRACTICE SESSION & PRELIM DONE.xlsx" file
and auto-creates Subject + Chapter records under the correct CourseLevel
(CSEET / CS Executive / CS Professional) of the main "Company Secretary"
course (CRS-0007).

Usage:
    python manage.py import_chapters                         # default path
    python manage.py import_chapters --file /path/to/file.xlsx
    python manage.py import_chapters --dry-run               # preview only
    python manage.py import_chapters --course-code CRS-0007  # explicit course
"""

import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
import openpyxl

from batches.models import Course, CourseLevel, Subject, Chapter


# ── Sheet-name → CourseLevel mapping ──────────────────────────────────────────
# Each sheet name contains a keyword that maps to a level name in the DB.
SHEET_LEVEL_MAP = {
    'PRO':   'CS Professional',
    'EXE':   'CS Executive',
    'CSEET': 'CSEET',
}

# Known subject abbreviation → full name overrides (optional, for clarity)
SUBJECT_NAME_OVERRIDES = {
    'ESG':    'ESG (Environment, Social & Governance)',
    'DRAFTING': 'Drafting, Appearances and Pleadings',
    'CMADD':  'Compliance Management, Audit and Due Diligence',
    'IPR':    'Intellectual Property Rights',
    'SMCF':   'Strategic Management and Corporate Finance',
    'CRVI':   'Corporate Restructuring, Valuation and Insolvency',
    'IBC':    'Insolvency and Bankruptcy Code',
    'JIGL':   'Jurisprudence, Interpretation and General Laws',
    'CLPC':   'Company Law Practice and Compliance',
    'SUBILL': 'Setting Up of Business and Industrial & Labour Laws',
    'CAFM':   'Corporate Accounting and Financial Management',
    'CMSL':   'Capital Markets and Securities Laws',
    'ECIPL':  'Economic, Commercial and Intellectual Property Laws',
    'TAX':    'Tax Laws and Practice',
}

DEFAULT_XLSX_PATH = '/home/nikita/Downloads/CHAPTER, CLASS TEST, PRACTICE SESSION & PRELIM DONE.xlsx'


class Command(BaseCommand):
    help = 'Import subjects and chapters from the chapters xlsx file into the database.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file', '-f',
            default=DEFAULT_XLSX_PATH,
            help='Path to the xlsx file (default: Downloads folder)',
        )
        parser.add_argument(
            '--course-code',
            default='CRS-0007',
            help='Course code to attach levels/subjects to (default: CRS-0007)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview what would be created without writing to DB',
        )

    def handle(self, *args, **options):
        xlsx_path = options['file']
        course_code = options['course_code']
        dry_run = options['dry_run']

        # ── Load workbook ─────────────────────────────────────────────────────
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {xlsx_path}')
        except Exception as e:
            raise CommandError(f'Error opening xlsx: {e}')

        # ── Resolve course ────────────────────────────────────────────────────
        try:
            course = Course.objects.get(code=course_code)
        except Course.DoesNotExist:
            raise CommandError(f'Course with code "{course_code}" not found.')

        self.stdout.write(f'\n📘 Course: {course.code} — {course.name}')
        self.stdout.write(f'📄 File:   {xlsx_path}')
        self.stdout.write(f'📋 Sheets: {wb.sheetnames}\n')

        if dry_run:
            self.stdout.write(self.style.WARNING('⚠️  DRY RUN — no changes will be saved.\n'))

        stats = {'subjects_created': 0, 'subjects_existing': 0,
                 'chapters_created': 0, 'chapters_existing': 0}

        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                level = self._resolve_level(sheet_name, course)
                if level is None:
                    self.stdout.write(self.style.WARNING(
                        f'  ⏭️  Skipping sheet "{sheet_name}" — cannot map to a CourseLevel.'
                    ))
                    continue

                self.stdout.write(self.style.SUCCESS(
                    f'\n━━━ Sheet: "{sheet_name}" → Level: {level.name} ━━━'
                ))

                ws = wb[sheet_name]
                self._process_sheet(ws, level, course, dry_run, stats)

            if dry_run:
                # Roll back in dry-run mode
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f'\n✅ Done! '
            f'Subjects: {stats["subjects_created"]} created, {stats["subjects_existing"]} already existed. '
            f'Chapters: {stats["chapters_created"]} created, {stats["chapters_existing"]} already existed.'
        ))

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _resolve_level(self, sheet_name, course):
        """Map a sheet name like "PRO JUNE'26" to the correct CourseLevel."""
        upper = sheet_name.upper()
        for keyword, level_name in SHEET_LEVEL_MAP.items():
            if keyword in upper:
                try:
                    return CourseLevel.objects.get(course=course, name=level_name)
                except CourseLevel.DoesNotExist:
                    self.stdout.write(self.style.WARNING(
                        f'  ⚠️  CourseLevel "{level_name}" not found under {course.code}. Skipping.'
                    ))
                    return None
        return None

    def _process_sheet(self, ws, level, course, dry_run, stats):
        """Parse one worksheet: detect subject headers and chapter rows."""
        current_subject = None
        chapter_order = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            col_a = row[0]
            if col_a is None or str(col_a).strip() == '':
                continue

            text = str(col_a).strip()

            # Skip the sheet title rows like "CS PROFESSIONAL JUNE 2026"
            if self._is_title_row(text):
                continue

            # Skip Part headers like "Part A: ..." or "Part B – ..."
            if self._is_part_header(text):
                continue

            # Detect subject header: a row where col_a has a short subject code/name
            # AND col_b (or nearby) has "CH. DONE" / "CH  DONE"
            col_b = row[1] if len(row) > 1 else None
            if col_b and 'CH' in str(col_b).upper() and 'DONE' in str(col_b).upper():
                # This is a subject header row
                subject_key = text.strip()
                subject_name = SUBJECT_NAME_OVERRIDES.get(subject_key, subject_key)

                current_subject, created = Subject.objects.get_or_create(
                    level=level,
                    name=subject_name,
                    defaults={
                        'organization': course.organization,
                    },
                )
                if created:
                    stats['subjects_created'] += 1
                    self.stdout.write(f'  ✨ Created subject: {current_subject.code} — {subject_name}')
                else:
                    stats['subjects_existing'] += 1
                    self.stdout.write(f'  📌 Existing subject: {current_subject.code} — {subject_name}')

                chapter_order = 0
                continue

            # Detect chapter row: starts with a number like "1. " or "1- " or "1 "
            if current_subject and self._is_chapter_row(text):
                chapter_order += 1
                chapter_name = self._clean_chapter_name(text)

                ch, created = Chapter.objects.get_or_create(
                    subject=current_subject,
                    order=chapter_order,
                    defaults={
                        'name': chapter_name,
                        'is_active': True,
                        'duration_hours': 0,
                    },
                )
                if created:
                    stats['chapters_created'] += 1
                    self.stdout.write(f'    📖 Ch {chapter_order}: {chapter_name}')
                else:
                    stats['chapters_existing'] += 1
                    # Update name if it changed in the xlsx
                    if ch.name != chapter_name:
                        ch.name = chapter_name
                        if not dry_run:
                            ch.save(update_fields=['name'])
                        self.stdout.write(f'    🔄 Ch {chapter_order}: {ch.name} → {chapter_name}')

    def _is_title_row(self, text):
        """Check if a row is a sheet title like 'CS PROFESSIONAL JUNE 2026'."""
        upper = text.upper()
        return any(kw in upper for kw in [
            'CS PROFESSIONAL', 'CS EXECUTIVE', 'DECEMBER 2026',
            'JUNE 2026', 'OCTOBER 2026',
        ])

    def _is_part_header(self, text):
        """Check if a row is a section part header like 'Part A: ...'."""
        stripped = text.strip()
        return bool(re.match(r'^Part\s+[A-Z]', stripped, re.IGNORECASE))

    def _is_chapter_row(self, text):
        """Check if text starts with a chapter number pattern like '1. ' or '1- '."""
        stripped = text.strip()
        return bool(re.match(r'^\d+[\.\-\)\s]', stripped))

    def _clean_chapter_name(self, text):
        """Remove the leading number prefix: '1. FOO' → 'FOO', '1- FOO' → 'FOO'."""
        cleaned = re.sub(r'^\d+[\.\-\)\s]+', '', text.strip())
        # Title case for consistency but preserve acronyms
        return cleaned.strip()
